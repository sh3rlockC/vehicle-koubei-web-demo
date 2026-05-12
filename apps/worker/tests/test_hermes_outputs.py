from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from worker_app.hermes_outputs import (
    _build_aggregate_prompt,
    _build_batch_prompt,
    _call_hermes,
    _extract_json,
    _runtime_provider,
    extract_whitelisted_comments,
    generate_outputs,
    generate_time_report_outputs,
)


def _write_input_workbooks(tmp_path: Path) -> tuple[Path, Path]:
    zj_path = tmp_path / "ZJ测试车原始口碑.xlsx"
    zj_workbook = Workbook()
    zj_sheet = zj_workbook.active
    zj_sheet.title = "购车口碑"
    zj_sheet.append(["用户名", "来源链接", "购车地", "发表日期", "最满意", "最不满意", "评价详情"])
    zj_sheet.append(["张三", "https://example.invalid/a", "上海浦东", "2026-03-01", "空间大", "内饰一般", "空间大，内饰一般"])
    zj_workbook.save(zj_path)

    dcd_path = tmp_path / "DCD口碑_测试车.xlsx"
    dcd_workbook = Workbook()
    dcd_sheet = dcd_workbook.active
    dcd_sheet.title = "口碑明细"
    dcd_sheet.append(["用户名", "来源链接", "购车城市", "发布时间", "评价全文"])
    dcd_sheet.append(["李四", "https://example.invalid/b", "北京朝阳", "2026-03-02", "动力顺，车机偶发卡顿"])
    dcd_workbook.save(dcd_path)

    return zj_path, dcd_path


def _write_time_range_workbooks(tmp_path: Path) -> tuple[Path, Path]:
    zj_path = tmp_path / "ZJ测试车原始口碑.xlsx"
    zj_workbook = Workbook()
    zj_sheet = zj_workbook.active
    zj_sheet.title = "购车口碑"
    zj_sheet.append(["用户名", "来源链接", "购车地", "发表日期", "最满意", "最不满意", "评价详情"])
    zj_sheet.append(["张三", "https://example.invalid/a", "上海浦东", "2026-03-01", "空间大", "内饰一般", "空间大，内饰一般"])
    zj_sheet.append(["王五", "https://example.invalid/c", "杭州西湖", "2026-03-03", "底盘稳", "胎噪大", "底盘稳，胎噪大"])
    zj_sheet.append(["赵六", "https://example.invalid/d", "广州天河", "", "配置高", "无", "无日期评论"])
    zj_workbook.save(zj_path)

    dcd_path = tmp_path / "DCD口碑_测试车.xlsx"
    dcd_workbook = Workbook()
    dcd_sheet = dcd_workbook.active
    dcd_sheet.title = "口碑明细"
    dcd_sheet.append(["用户名", "来源链接", "购车城市", "发布时间", "评价全文"])
    dcd_sheet.append(["李四", "https://example.invalid/b", "北京朝阳", "2026-03-02", "动力顺，车机偶发卡顿"])
    dcd_sheet.append(["孙七", "https://example.invalid/e", "深圳南山", "2026-03-04", "空间不错，续航一般"])
    dcd_workbook.save(dcd_path)

    return zj_path, dcd_path


def test_extract_whitelisted_comments_removes_private_fields(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_input_workbooks(tmp_path)

    comments = extract_whitelisted_comments(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        model_name="测试车",
    )

    serialized = json.dumps(comments, ensure_ascii=False)
    assert "张三" not in serialized
    assert "李四" not in serialized
    assert "example.invalid" not in serialized
    assert "上海浦东" not in serialized
    assert "北京朝阳" not in serialized
    assert comments[0]["platform"] == "汽车之家"
    assert comments[0]["positive_text"] == "空间大"
    assert comments[0]["full_text"] == "空间大，内饰一般"
    assert comments[1]["platform"] == "懂车帝"
    assert comments[1]["full_text"] == "动力顺，车机偶发卡顿"


def test_extract_whitelisted_comments_builds_standard_raw_comment_json(tmp_path: Path) -> None:
    zj_path = tmp_path / "ZJ测试车原始口碑.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "购车口碑"
    sheet.append(["用户名", "来源链接", "购车地", "发表日期", "评价详情"])
    sheet.append(
        [
            "张三",
            "https://example.invalid/a",
            "上海浦东",
            "2026-03-01",
            "最满意：空间大，二排宽敞。 最不满意：车机偶发卡顿。 空间：第三排短途够用。 "
            "驾驶感受：底盘稳。 续航：纯电通勤够用。 外观：大气。 内饰：用料不错。 "
            "性价比：配置高。 智能化：语音识别准确。",
        ]
    )
    workbook.save(zj_path)

    comments = extract_whitelisted_comments(
        autohome_input=zj_path,
        dcd_input=None,
        model_name="测试车",
    )

    assert len(comments) == 1
    comment = comments[0]
    assert comment["positive_text"] == "空间大，二排宽敞。"
    assert comment["negative_text"] == "车机偶发卡顿。"
    assert comment["raw_comment"]["positive_text"] == "空间大，二排宽敞。"
    assert comment["raw_comment"]["negative_text"] == "车机偶发卡顿。"
    assert comment["raw_comment"]["sections"] == {
        "space": "第三排短途够用。",
        "driving": "底盘稳。",
        "range": "纯电通勤够用。",
        "appearance": "大气。",
        "interior": "用料不错。",
        "cost_value": "配置高。",
        "intelligence": "语音识别准确。",
    }
    serialized = json.dumps(comment, ensure_ascii=False)
    assert "张三" not in serialized
    assert "example.invalid" not in serialized
    assert "上海浦东" not in serialized


def test_batch_prompt_uses_standard_raw_comment_json_shape() -> None:
    comments = [
        {
            "comment_id": "autohome_0001",
            "platform": "汽车之家",
            "date": "2026-03-01",
            "model_name": "测试车",
            "positive_text": "空间大",
            "negative_text": "车机卡顿",
            "full_text": "最满意：空间大。最不满意：车机卡顿。空间：二排宽敞。",
            "raw_comment": {
                "positive_text": "空间大",
                "negative_text": "车机卡顿",
                "full_text": "最满意：空间大。最不满意：车机卡顿。空间：二排宽敞。",
                "sections": {"space": "二排宽敞。"},
            },
        }
    ]

    prompt = _build_batch_prompt(model_name="测试车", batch_index=1, total_batches=1, comments=comments)

    assert '"raw_comment"' in prompt
    assert '"sections": {"space": "二排宽敞。"}' in prompt
    assert '"comment_id": "autohome_0001"' in prompt
    assert '"source_url"' not in prompt
    assert '"purchase_location"' not in prompt


def test_runtime_provider_prefers_deepseek_provider_even_with_base_url() -> None:
    provider, model, env = _runtime_provider(
        {
            "LLM_PROVIDER": "deepseek",
            "LLM_MODEL_REPORT": "deepseekv4pro",
            "LLM_API_KEY": "test-key",
            "LLM_BASE_URL": "https://api.deepseek.com",
        }
    )

    assert provider == "deepseek"
    assert model == "deepseekv4pro"
    assert env["DEEPSEEK_API_KEY"] == "test-key"


def test_aggregate_prompt_compacts_large_batch_payloads() -> None:
    comments = [{"platform": "汽车之家"}, {"platform": "懂车帝"}]
    long_summary = "长摘要" * 300
    batch_payloads = [
        {
            "themes": [
                {
                    "direction": "positive",
                    "term": f"主题{index}",
                    "count": 50 - index,
                    "summary": long_summary,
                    "description": long_summary,
                    "evidence_ids": [f"comment_{item:04d}" for item in range(20)],
                    "unused_extra": "不应进入汇总prompt" * 50,
                }
                for index in range(20)
            ],
            "suggestions": [{"direction": "产品", "text": "建议" * 200} for _ in range(20)],
            "platform_notes": [{"platform": "汽车之家", "summary": "平台差异" * 200} for _ in range(8)],
            "boss_brief": ["老板汇报" * 200 for _ in range(6)],
        }
        for _ in range(4)
    ]

    prompt = _build_aggregate_prompt(model_name="测试车", comments=comments, batch_payloads=batch_payloads)

    assert len(prompt) < 70_000
    assert "comment_0003" not in prompt
    assert "不应进入汇总prompt" not in prompt
    assert "长摘要" * 120 not in prompt


def test_aggregate_prompt_stays_below_command_argument_limit_for_many_batches() -> None:
    comments = [{"platform": "汽车之家"} for _ in range(1000)] + [{"platform": "懂车帝"} for _ in range(142)]
    long_summary = "长摘要" * 120
    batch_payloads = [
        {
            "batch": f"{index}/60",
            "themes": [
                {
                    "direction": "positive" if theme_index % 2 == 0 else "negative",
                    "term": f"主题{index}_{theme_index}",
                    "count": 100 - theme_index,
                    "summary": long_summary,
                    "evidence_ids": [f"comment_{index}_{item}" for item in range(20)],
                }
                for theme_index in range(16)
            ],
            "suggestions": [{"direction": "产品", "text": "建议" * 160} for _ in range(8)],
            "platform_notes": [{"platform": "汽车之家", "summary": "平台差异" * 160} for _ in range(4)],
            "boss_brief": ["老板汇报" * 160 for _ in range(4)],
        }
        for index in range(60)
    ]

    prompt = _build_aggregate_prompt(model_name="风云T11", comments=comments, batch_payloads=batch_payloads)

    assert len(prompt.encode("utf-8")) < 90_000
    assert "长摘要" * 60 not in prompt
    assert "comment_1_3" not in prompt


def test_extract_json_cleans_hermes_warning_fence_and_inner_quotes() -> None:
    raw_response = (
        "⚠️ Normalized model 'deepseekv4pro' to 'deepseek-chat' for deepseek.\n"
        "```json\n"
        '{"headline":"风云T11空间好评突出",'
        '"weakness_blocks":[{"title":"第三排舒适性","summary":"用户认为第三排像"坐小板凳"，'
        '外观像"小揽胜"，需要优化。","evidence_ids":["autohome_0001"]}]}\n'
        "```"
    )

    payload = _extract_json(raw_response)

    assert payload["headline"] == "风云T11空间好评突出"
    assert payload["weakness_blocks"][0]["summary"] == '用户认为第三排像"坐小板凳"，外观像"小揽胜"，需要优化。'


def test_call_hermes_timeout_uses_concise_error_without_prompt(tmp_path: Path) -> None:
    fake_hermes = tmp_path / "hermes"
    fake_hermes.write_text("#!/bin/sh\nsleep 2\n", encoding="utf-8")
    fake_hermes.chmod(0o755)

    with pytest.raises(RuntimeError) as exc_info:
        _call_hermes(
            "SECRET_PROMPT_SHOULD_NOT_APPEAR",
            hermes_command=str(fake_hermes),
            env={**os.environ, "LLM_API_KEY": "test-key", "LLM_MODEL_REPORT": "deepseek-chat", "HERMES_TIMEOUT_SECONDS": "1"},
            call_label="aggregate",
        )

    message = str(exc_info.value)
    assert message == "hermes_timeout:aggregate:1s"
    assert "SECRET_PROMPT_SHOULD_NOT_APPEAR" not in message


def _write_fake_summary_script(path: Path) -> None:
    path.write_text(
        """
import argparse
import json
from pathlib import Path
from openpyxl import Workbook

parser = argparse.ArgumentParser()
parser.add_argument("--autohome-input")
parser.add_argument("--dcd-input")
parser.add_argument("--input")
parser.add_argument("--output", required=True)
parser.add_argument("--model-name", required=True)
parser.add_argument("--progress-file", required=True)
args = parser.parse_args()

workbook = Workbook()
workbook.active.title = "总览摘要"
workbook["总览摘要"].append(["模块", "内容"])
workbook["总览摘要"].append(["平台样本", "汽车之家 1 条；懂车帝 1 条"])
workbook["总览摘要"].append(["综合一句话", "Hermes 失败后规则兜底"])
workbook.create_sheet("跨平台对比").append(["方向", "汽车之家_优势提及", "汽车之家_槽点提及", "懂车帝_优势提及", "懂车帝_槽点提及"])
business = workbook.create_sheet("综合业务摘要")
business.append(["模块", "内容"])
business.append(["核心好评", "空间大"])
business.append(["核心槽点", "车机卡顿"])
opportunity = workbook.create_sheet("产品机会点")
opportunity.append(["类型", "方向", "建议"])
opportunity.append(["改进", "车机", "提升稳定性"])
one_pager = workbook.create_sheet("一页纸总结")
one_pager.append(["双平台口碑一页纸总结"])
one_pager.append(["最满意TOP5", "空间大"])
one_pager.append(["最不满意TOP5", "车机卡顿"])
Path(args.output).parent.mkdir(parents=True, exist_ok=True)
workbook.save(args.output)
Path(args.output).with_suffix(".validation.json").write_text("{}", encoding="utf-8")
Path(args.progress_file).write_text(json.dumps({"percent": 100}), encoding="utf-8")
Path(args.output).with_suffix(".args.json").write_text(json.dumps(vars(args), ensure_ascii=False), encoding="utf-8")
""",
        encoding="utf-8",
    )


def _write_fake_wordcloud_script(path: Path) -> None:
    path.write_text(
        """
import argparse
import json
from pathlib import Path
from openpyxl import Workbook

parser = argparse.ArgumentParser()
parser.add_argument("--input", required=True)
parser.add_argument("--output-dir", required=True)
parser.add_argument("--model-name", required=True)
parser.add_argument("--font-path")
parser.add_argument("--json", action="store_true")
args = parser.parse_args()

output_dir = Path(args.output_dir)
output_dir.mkdir(parents=True, exist_ok=True)
terms = output_dir / f"{args.model_name}_词云词项清单.xlsx"
workbook = Workbook()
workbook.active.title = "positive_terms"
workbook["positive_terms"].append(["term", "weight"])
workbook["positive_terms"].append(["空间", 2])
negative = workbook.create_sheet("negative_terms")
negative.append(["term", "weight"])
negative.append(["车机", 1])
workbook.save(terms)
images = []
for suffix in ["优点词云", "槽点词云"]:
    image = output_dir / f"{args.model_name}_{suffix}.png"
    image.write_bytes(b"\\x89PNG\\r\\n\\x1a\\n")
    images.append(str(image))
print(json.dumps({"excel_path": str(terms), "image_paths": images}, ensure_ascii=False))
""",
        encoding="utf-8",
    )


def test_generate_outputs_falls_back_when_hermes_json_is_invalid(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_input_workbooks(tmp_path)
    summary_script = tmp_path / "summary.py"
    wordcloud_script = tmp_path / "wordcloud.py"
    fake_hermes = tmp_path / "hermes"
    _write_fake_summary_script(summary_script)
    _write_fake_wordcloud_script(wordcloud_script)
    fake_hermes.write_text("#!/bin/sh\necho not-json\n", encoding="utf-8")
    fake_hermes.chmod(0o755)

    result = generate_outputs(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        postprocess_input=tmp_path / "dual.xlsx",
        summary_output=tmp_path / "summary" / "测试车_双平台口碑摘要.xlsx",
        terms_output=tmp_path / "wordcloud" / "测试车_词云词项清单.xlsx",
        wordcloud_output_dir=tmp_path / "wordcloud",
        final_report_output=tmp_path / "ai" / "final_report.json",
        qa_chunks_output=tmp_path / "ai" / "qa_chunks.json",
        model_name="测试车",
        progress_file=tmp_path / "progress" / "generating_hermes_outputs.progress.json",
        summary_script=summary_script,
        wordcloud_script=wordcloud_script,
        hermes_command=str(fake_hermes),
        env={**os.environ, "LLM_API_KEY": "test-key", "LLM_MODEL_REPORT": "deepseek-chat"},
    )

    assert result["status"] == "degraded"
    assert result["degraded"] is True
    assert result["fallback_reason"].startswith("hermes_invalid_json")
    assert Path(result["summary_path"]).exists()
    assert Path(result["terms_path"]).exists()
    assert Path(result["final_report_path"]).exists()
    assert Path(result["qa_chunks_path"]).exists()
    assert load_workbook(result["summary_path"], data_only=True)["综合业务摘要"]["A2"].value == "核心好评"


def test_generate_outputs_rule_fallback_uses_dcd_input_when_autohome_is_missing(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_input_workbooks(tmp_path)
    zj_path.unlink()
    summary_script = tmp_path / "summary.py"
    wordcloud_script = tmp_path / "wordcloud.py"
    fake_hermes = tmp_path / "hermes"
    _write_fake_summary_script(summary_script)
    _write_fake_wordcloud_script(wordcloud_script)
    fake_hermes.write_text("#!/bin/sh\necho not-json\n", encoding="utf-8")
    fake_hermes.chmod(0o755)
    summary_output = tmp_path / "summary" / "测试车_双平台口碑摘要.xlsx"

    result = generate_outputs(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        postprocess_input=None,
        summary_output=summary_output,
        terms_output=tmp_path / "wordcloud" / "测试车_词云词项清单.xlsx",
        wordcloud_output_dir=tmp_path / "wordcloud",
        final_report_output=tmp_path / "ai" / "final_report.json",
        qa_chunks_output=tmp_path / "ai" / "qa_chunks.json",
        model_name="测试车",
        progress_file=tmp_path / "progress" / "generating_hermes_outputs.progress.json",
        summary_script=summary_script,
        wordcloud_script=wordcloud_script,
        hermes_command=str(fake_hermes),
        single_platform=True,
        env={**os.environ, "LLM_API_KEY": "test-key", "LLM_MODEL_REPORT": "deepseek-chat"},
    )

    summary_args = json.loads(summary_output.with_suffix(".args.json").read_text(encoding="utf-8"))
    assert result["status"] == "degraded"
    assert summary_args["input"] == str(dcd_path)
    assert summary_args["autohome_input"] is None
    assert summary_args["dcd_input"] is None


def test_generate_outputs_repairs_invalid_hermes_json_before_fallback(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_input_workbooks(tmp_path)
    summary_script = tmp_path / "summary.py"
    wordcloud_script = tmp_path / "wordcloud.py"
    fake_hermes = tmp_path / "hermes"
    _write_fake_summary_script(summary_script)
    _write_fake_wordcloud_script(wordcloud_script)
    fake_hermes.write_text(
        """#!/usr/bin/env python3
import json
import sys
from pathlib import Path

state = Path(__file__).with_suffix(".state")
call_count = int(state.read_text(encoding="utf-8")) if state.exists() else 0
state.write_text(str(call_count + 1), encoding="utf-8")
prompt = sys.argv[-1]

if call_count == 0:
    print('{"themes":[{"direction":"positive" "term":"空间","count":1}],"suggestions":[]}')
elif call_count == 1:
    if '"direction":"positive" "term"' not in prompt:
        print('{"themes":[{"direction":"positive" "term":"空间","count":1}],"suggestions":[]}')
    else:
        print(json.dumps({
            "themes": [
                {"direction": "positive", "term": "空间", "count": 2, "summary": "空间大", "evidence_ids": ["autohome_0001"]},
                {"direction": "negative", "term": "车机", "count": 1, "summary": "车机卡顿", "evidence_ids": ["dcd_0001"]}
            ],
            "suggestions": [{"direction": "negative", "text": "优化车机稳定性"}],
            "platform_notes": [],
            "boss_brief": ["空间好评突出"]
        }, ensure_ascii=False))
else:
    print(json.dumps({
        "headline": "测试车空间好评突出，车机仍需优化。",
        "executive_summary": "Hermes JSON 修复后生成最终摘要。",
        "strength_blocks": [{"title": "核心好评", "summary": "空间表现突出", "evidence_ids": ["autohome_0001"]}],
        "weakness_blocks": [{"title": "核心槽点", "summary": "车机偶发卡顿", "evidence_ids": ["dcd_0001"]}],
        "platform_difference_blocks": [],
        "action_blocks": [{"title": "车机优化", "summary": "提升车机稳定性", "evidence_ids": ["dcd_0001"]}],
        "boss_brief": ["空间好评突出", "车机仍需优化"],
        "keyword_rankings": {
            "positive": [{"term": "空间", "count": 2}],
            "negative": [{"term": "车机", "count": 1}]
        },
        "qa_chunks": [],
        "compare_rows": [],
        "opportunity_rows": []
    }, ensure_ascii=False))
""",
        encoding="utf-8",
    )
    fake_hermes.chmod(0o755)

    result = generate_outputs(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        postprocess_input=tmp_path / "dual.xlsx",
        summary_output=tmp_path / "summary" / "测试车_双平台口碑摘要.xlsx",
        terms_output=tmp_path / "wordcloud" / "测试车_词云词项清单.xlsx",
        wordcloud_output_dir=tmp_path / "wordcloud",
        final_report_output=tmp_path / "ai" / "final_report.json",
        qa_chunks_output=tmp_path / "ai" / "qa_chunks.json",
        model_name="测试车",
        progress_file=tmp_path / "progress" / "generating_hermes_outputs.progress.json",
        summary_script=summary_script,
        wordcloud_script=wordcloud_script,
        hermes_command=str(fake_hermes),
        env={**os.environ, "LLM_API_KEY": "test-key", "LLM_MODEL_REPORT": "deepseek-chat", "HERMES_JSON_RETRIES": "1"},
    )

    assert result["status"] == "success"
    assert result["degraded"] is False
    assert json.loads(Path(result["final_report_path"]).read_text(encoding="utf-8"))["headline"] == "测试车空间好评突出，车机仍需优化。"
    normalized_comments_path = Path(result["normalized_comments_path"])
    assert normalized_comments_path.exists()
    normalized_comment = json.loads(normalized_comments_path.read_text(encoding="utf-8").splitlines()[0])
    assert normalized_comment["comment_id"] == "autohome_0001"
    assert "raw_comment" in normalized_comment
    assert "sections" in normalized_comment["raw_comment"]
    assert (tmp_path / "logs" / "hermes" / "batch_001.attempt_1.stdout.txt").exists()
    assert (tmp_path / "logs" / "hermes" / "batch_001.attempt_1.parse_error.txt").exists()


def test_generate_outputs_uses_local_aggregate_when_hermes_aggregate_times_out(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_input_workbooks(tmp_path)
    summary_script = tmp_path / "summary.py"
    wordcloud_script = tmp_path / "wordcloud.py"
    fake_hermes = tmp_path / "hermes"
    _write_fake_summary_script(summary_script)
    _write_fake_wordcloud_script(wordcloud_script)
    fake_hermes.write_text(
        """#!/usr/bin/env python3
import json
import sys
import time

prompt = sys.argv[-1]
if "请归并各批分析结果" in prompt:
    time.sleep(2)
else:
    print(json.dumps({
        "themes": [
            {"direction": "positive", "term": "空间", "count": 2, "summary": "空间大", "evidence_ids": ["autohome_0001"]},
            {"direction": "negative", "term": "车机", "count": 1, "summary": "车机卡顿", "evidence_ids": ["dcd_0001"]}
        ],
        "suggestions": [{"direction": "产品", "text": "优化车机稳定性"}],
        "platform_notes": [{"platform": "懂车帝", "summary": "更关注车机体验"}],
        "boss_brief": ["空间好评突出"]
    }, ensure_ascii=False))
""",
        encoding="utf-8",
    )
    fake_hermes.chmod(0o755)

    result = generate_outputs(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        postprocess_input=tmp_path / "dual.xlsx",
        summary_output=tmp_path / "summary" / "测试车_双平台口碑摘要.xlsx",
        terms_output=tmp_path / "wordcloud" / "测试车_词云词项清单.xlsx",
        wordcloud_output_dir=tmp_path / "wordcloud",
        final_report_output=tmp_path / "ai" / "final_report.json",
        qa_chunks_output=tmp_path / "ai" / "qa_chunks.json",
        model_name="测试车",
        progress_file=tmp_path / "progress" / "generating_hermes_outputs.progress.json",
        summary_script=summary_script,
        wordcloud_script=wordcloud_script,
        hermes_command=str(fake_hermes),
        env={
            **os.environ,
            "LLM_API_KEY": "test-key",
            "LLM_MODEL_REPORT": "deepseek-chat",
            "HERMES_TIMEOUT_SECONDS": "1",
        },
    )

    assert result["status"] == "success"
    assert result["degraded"] is False
    assert result["source"] == "hermes-local-aggregate"
    assert result["aggregate_fallback_reason"] == "hermes_timeout:aggregate:1s"
    assert (tmp_path / "logs" / "hermes" / "aggregate.attempt_1.timeout.txt").exists()
    report = json.loads(Path(result["final_report_path"]).read_text(encoding="utf-8"))
    assert "空间" in report["headline"]
    assert "车机" in report["headline"]


def test_generate_outputs_uses_local_batch_when_one_batch_json_remains_invalid(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_input_workbooks(tmp_path)
    summary_script = tmp_path / "summary.py"
    wordcloud_script = tmp_path / "wordcloud.py"
    fake_hermes = tmp_path / "hermes"
    state_path = tmp_path / "fake-hermes-count.txt"
    _write_fake_summary_script(summary_script)
    _write_fake_wordcloud_script(wordcloud_script)
    fake_hermes.write_text(
        """#!/usr/bin/env python3
import json
import sys
from pathlib import Path

state_path = Path("__STATE_PATH__")
count = int(state_path.read_text(encoding="utf-8")) if state_path.exists() else 0
count += 1
state_path.write_text(str(count), encoding="utf-8")
prompt = sys.argv[-1]
if "请归并各批分析结果" in prompt:
    assert "本地批次兜底" in prompt
    print(json.dumps({
        "headline": "测试车全量报告已生成。",
        "executive_summary": "单批 Hermes JSON 失败后使用本地批次兜底继续完成。",
        "strength_blocks": [{"title": "核心好评", "summary": "空间大、本地批次兜底", "evidence_ids": ["autohome_0001"]}],
        "weakness_blocks": [{"title": "核心槽点", "summary": "车机卡顿、本地槽点", "evidence_ids": ["dcd_0001"]}],
        "platform_difference_blocks": [],
        "action_blocks": [{"title": "产品建议", "summary": "复核本地兜底批次并优化车机", "evidence_ids": ["dcd_0001"]}],
        "boss_brief": ["全量报告完成"],
        "keyword_rankings": {
            "positive": [{"term": "空间", "count": 1}, {"term": "本地批次兜底", "count": 1}],
            "negative": [{"term": "车机", "count": 1}, {"term": "本地槽点", "count": 1}]
        },
        "qa_chunks": [],
        "compare_rows": [],
        "opportunity_rows": []
    }, ensure_ascii=False))
elif count == 1:
    print(json.dumps({
        "themes": [
            {"direction": "positive", "term": "空间", "count": 1, "summary": "空间大", "evidence_ids": ["autohome_0001"]},
            {"direction": "negative", "term": "内饰", "count": 1, "summary": "内饰一般", "evidence_ids": ["autohome_0001"]}
        ],
        "suggestions": [],
        "platform_notes": [],
        "boss_brief": ["空间好评突出"]
    }, ensure_ascii=False))
else:
    print('{"themes":[{"direction":"negative","term":"车机","count":1}')
""".replace("__STATE_PATH__", str(state_path)),
        encoding="utf-8",
    )
    fake_hermes.chmod(0o755)

    result = generate_outputs(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        postprocess_input=tmp_path / "dual.xlsx",
        summary_output=tmp_path / "summary" / "测试车_双平台口碑摘要.xlsx",
        terms_output=tmp_path / "wordcloud" / "测试车_词云词项清单.xlsx",
        wordcloud_output_dir=tmp_path / "wordcloud",
        final_report_output=tmp_path / "ai" / "final_report.json",
        qa_chunks_output=tmp_path / "ai" / "qa_chunks.json",
        model_name="测试车",
        progress_file=tmp_path / "progress" / "generating_hermes_outputs.progress.json",
        summary_script=summary_script,
        wordcloud_script=wordcloud_script,
        hermes_command=str(fake_hermes),
        env={
            **os.environ,
            "LLM_API_KEY": "test-key",
            "LLM_MODEL_REPORT": "deepseek-chat",
            "HERMES_BATCH_SIZE": "1",
            "HERMES_JSON_RETRIES": "1",
        },
    )

    assert result["status"] == "success"
    assert result["degraded"] is False
    assert result["source"] == "hermes-partial-local-batch"
    assert result["batch_fallbacks"] == [{"batch": 2, "reason": "hermes_invalid_json:response JSON was not balanced"}]
    assert Path(result["final_report_path"]).exists()
    report = json.loads(Path(result["final_report_path"]).read_text(encoding="utf-8"))
    assert report["headline"] == "测试车全量报告已生成。"


def test_generate_outputs_clears_stale_hermes_debug_logs(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_input_workbooks(tmp_path)
    summary_script = tmp_path / "summary.py"
    wordcloud_script = tmp_path / "wordcloud.py"
    fake_hermes = tmp_path / "hermes"
    stale_log = tmp_path / "logs" / "hermes" / "batch_999.attempt_1.parse_error.txt"
    stale_log.parent.mkdir(parents=True)
    stale_log.write_text("old parse error", encoding="utf-8")
    _write_fake_summary_script(summary_script)
    _write_fake_wordcloud_script(wordcloud_script)
    fake_hermes.write_text(
        """#!/usr/bin/env python3
import json

print(json.dumps({
    "headline": "测试车空间好评突出，车机仍需优化。",
    "executive_summary": "Hermes 生成摘要。",
    "strength_blocks": [{"title": "核心好评", "summary": "空间表现突出", "evidence_ids": ["autohome_0001"]}],
    "weakness_blocks": [{"title": "核心槽点", "summary": "车机偶发卡顿", "evidence_ids": ["dcd_0001"]}],
    "platform_difference_blocks": [],
    "action_blocks": [{"title": "车机优化", "summary": "提升车机稳定性", "evidence_ids": ["dcd_0001"]}],
    "boss_brief": ["空间好评突出"],
    "keyword_rankings": {
        "positive": [{"term": "空间", "count": 2}],
        "negative": [{"term": "车机", "count": 1}]
    },
    "qa_chunks": [],
    "compare_rows": [],
    "opportunity_rows": []
}, ensure_ascii=False))
""",
        encoding="utf-8",
    )
    fake_hermes.chmod(0o755)

    result = generate_outputs(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        postprocess_input=tmp_path / "dual.xlsx",
        summary_output=tmp_path / "summary" / "测试车_双平台口碑摘要.xlsx",
        terms_output=tmp_path / "wordcloud" / "测试车_词云词项清单.xlsx",
        wordcloud_output_dir=tmp_path / "wordcloud",
        final_report_output=tmp_path / "ai" / "final_report.json",
        qa_chunks_output=tmp_path / "ai" / "qa_chunks.json",
        model_name="测试车",
        progress_file=tmp_path / "progress" / "generating_hermes_outputs.progress.json",
        summary_script=summary_script,
        wordcloud_script=wordcloud_script,
        hermes_command=str(fake_hermes),
        env={**os.environ, "LLM_API_KEY": "test-key", "LLM_MODEL_REPORT": "deepseek-chat"},
    )

    assert result["status"] == "success"
    assert not stale_log.exists()


def test_generate_time_report_outputs_filters_comments_by_inclusive_date_range(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_time_range_workbooks(tmp_path)
    fake_hermes = tmp_path / "hermes"
    fake_hermes.write_text(
        """#!/usr/bin/env python3
import json
import sys

prompt = sys.argv[-1]
if "请归并各批分析结果" in prompt:
    assert "空间大，内饰一般" not in prompt
    print(json.dumps({
        "headline": "测试车在 2026-03-02 至 2026-03-03 的好评集中在动力和底盘，槽点集中在车机和胎噪。",
        "executive_summary": "按时间范围筛选后的 Hermes 一页纸。",
        "strength_blocks": [{"title": "核心好评", "summary": "动力顺、底盘稳", "evidence_ids": ["dcd_0001", "autohome_0002"]}],
        "weakness_blocks": [{"title": "核心槽点", "summary": "车机偶发卡顿、胎噪大", "evidence_ids": ["dcd_0001", "autohome_0002"]}],
        "platform_difference_blocks": [],
        "action_blocks": [{"title": "产品建议", "summary": "优化车机和NVH", "evidence_ids": ["dcd_0001"]}],
        "boss_brief": ["时间范围内动力和底盘好评突出"],
        "keyword_rankings": {
            "positive": [{"term": "动力", "count": 1}, {"term": "底盘", "count": 1}],
            "negative": [{"term": "车机", "count": 1}, {"term": "胎噪", "count": 1}]
        },
        "qa_chunks": [],
        "compare_rows": [],
        "opportunity_rows": []
    }, ensure_ascii=False))
else:
    assert "2026-03-01" not in prompt
    assert "2026-03-04" not in prompt
    assert "无日期评论" not in prompt
    print(json.dumps({
        "themes": [
            {"direction": "positive", "term": "动力", "count": 1, "summary": "动力顺", "evidence_ids": ["dcd_0001"]},
            {"direction": "positive", "term": "底盘", "count": 1, "summary": "底盘稳", "evidence_ids": ["autohome_0002"]},
            {"direction": "negative", "term": "车机", "count": 1, "summary": "车机偶发卡顿", "evidence_ids": ["dcd_0001"]},
            {"direction": "negative", "term": "胎噪", "count": 1, "summary": "胎噪大", "evidence_ids": ["autohome_0002"]}
        ],
        "suggestions": [{"direction": "产品", "text": "优化车机和NVH"}],
        "platform_notes": [],
        "boss_brief": ["动力和底盘好评突出"]
    }, ensure_ascii=False))
""",
        encoding="utf-8",
    )
    fake_hermes.chmod(0o755)

    result = generate_time_report_outputs(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        output_dir=tmp_path / "time_report",
        model_name="测试车",
        start_date="2026-03-02",
        end_date="2026-03-03",
        hermes_command=str(fake_hermes),
        env={**os.environ, "LLM_API_KEY": "test-key", "LLM_MODEL_REPORT": "deepseek-chat", "HERMES_BATCH_SIZE": "1"},
    )

    assert result["status"] == "completed"
    assert result["sample_count"] == 2
    assert result["platform_counts"] == {"汽车之家": 1, "懂车帝": 1}
    assert Path(result["summary_path"]).exists()
    assert Path(result["terms_path"]).exists()
    assert Path(result["final_report_path"]).exists()
    assert Path(result["qa_chunks_path"]).exists()
    assert len(result["image_paths"]) == 2
    report = json.loads(Path(result["final_report_path"]).read_text(encoding="utf-8"))
    assert "2026-03-02 至 2026-03-03" in report["headline"]


def test_generate_time_report_outputs_rejects_empty_date_range(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_time_range_workbooks(tmp_path)

    with pytest.raises(ValueError) as exc_info:
        generate_time_report_outputs(
            autohome_input=zj_path,
            dcd_input=dcd_path,
            output_dir=tmp_path / "time_report",
            model_name="测试车",
            start_date="2026-04-01",
            end_date="2026-04-02",
            hermes_command=str(tmp_path / "missing-hermes"),
            env={**os.environ, "LLM_API_KEY": "test-key"},
        )

    assert str(exc_info.value) == "no_comments_in_date_range"


def test_generate_time_report_outputs_uses_local_batch_when_one_batch_times_out(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_time_range_workbooks(tmp_path)
    fake_hermes = tmp_path / "hermes"
    state_path = tmp_path / "fake-hermes-count.txt"
    fake_hermes.write_text(
        """#!/usr/bin/env python3
import json
import sys
import time
from pathlib import Path

state_path = Path("__STATE_PATH__")
count = int(state_path.read_text(encoding="utf-8")) if state_path.exists() else 0
count += 1
state_path.write_text(str(count), encoding="utf-8")
prompt = sys.argv[-1]
if "请归并各批分析结果" in prompt:
    assert "空间大，内饰一般" not in prompt
    assert "本地批次兜底" in prompt
    print(json.dumps({
        "headline": "测试车时间范围报告已生成。",
        "executive_summary": "单批 Hermes 超时后使用本地批次兜底继续完成。",
        "strength_blocks": [{"title": "核心好评", "summary": "动力顺、底盘稳", "evidence_ids": ["dcd_0001", "autohome_0002"]}],
        "weakness_blocks": [{"title": "核心槽点", "summary": "车机卡顿、胎噪大", "evidence_ids": ["dcd_0001", "autohome_0002"]}],
        "platform_difference_blocks": [],
        "action_blocks": [{"title": "产品建议", "summary": "继续跟进批次兜底识别出的槽点", "evidence_ids": ["autohome_0002"]}],
        "boss_brief": ["时间范围报告完成"],
        "keyword_rankings": {
            "positive": [{"term": "动力", "count": 1}, {"term": "本地批次兜底", "count": 1}],
            "negative": [{"term": "车机", "count": 1}, {"term": "本地槽点", "count": 1}]
        },
        "qa_chunks": [],
        "compare_rows": [],
        "opportunity_rows": []
    }, ensure_ascii=False))
elif count == 2:
    time.sleep(2)
else:
    print(json.dumps({
        "themes": [
            {"direction": "positive", "term": "动力", "count": 1, "summary": "动力顺", "evidence_ids": ["dcd_0001"]},
            {"direction": "negative", "term": "车机", "count": 1, "summary": "车机偶发卡顿", "evidence_ids": ["dcd_0001"]}
        ],
        "suggestions": [],
        "platform_notes": [],
        "boss_brief": ["动力好评突出"]
    }, ensure_ascii=False))
""".replace("__STATE_PATH__", str(state_path)),
        encoding="utf-8",
    )
    fake_hermes.chmod(0o755)

    result = generate_time_report_outputs(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        output_dir=tmp_path / "time_report",
        model_name="测试车",
        start_date="2026-03-02",
        end_date="2026-03-03",
        hermes_command=str(fake_hermes),
        env={
            **os.environ,
            "LLM_API_KEY": "test-key",
            "LLM_MODEL_REPORT": "deepseek-chat",
            "HERMES_BATCH_SIZE": "1",
            "HERMES_TIMEOUT_SECONDS": "1",
        },
    )

    assert result["status"] == "completed"
    assert result["source"] == "hermes-partial-local-batch"
    assert result["batch_fallbacks"] == [{"batch": 2, "reason": "hermes_timeout:batch_002:1s"}]
    assert Path(result["final_report_path"]).exists()
    assert json.loads(Path(result["final_report_path"]).read_text(encoding="utf-8"))["sample_count"] == 2


def test_generate_outputs_normalizes_legacy_hermes_labels(tmp_path: Path) -> None:
    zj_path, dcd_path = _write_input_workbooks(tmp_path)
    summary_script = tmp_path / "summary.py"
    wordcloud_script = tmp_path / "wordcloud.py"
    fake_hermes = tmp_path / "hermes"
    _write_fake_summary_script(summary_script)
    _write_fake_wordcloud_script(wordcloud_script)
    fake_hermes.write_text(
        """#!/usr/bin/env python3
import json

print(json.dumps({
    "headline": "核心卖点集中在空间，核心槽点TOP是车机",
    "executive_summary": "核心卖点TOP需要保留传播，核心槽点TOP需要产品跟进",
    "strength_blocks": [{"title": "核心卖点", "summary": "核心卖点TOP为空间", "evidence_ids": ["old.strength"]}],
    "weakness_blocks": [{"title": "核心槽点TOP", "summary": "核心槽点TOP为车机", "evidence_ids": ["old.weakness"]}],
    "platform_difference_blocks": [],
    "action_blocks": [{"title": "核心槽点TOP治理", "summary": "围绕核心槽点TOP改进", "evidence_ids": ["old.action"]}],
    "boss_brief": ["核心卖点TOP讲空间", "核心槽点TOP讲车机"],
    "keyword_rankings": {
        "positive": [{"term": "空间", "count": 2}],
        "negative": [{"term": "车机", "count": 1}]
    },
    "qa_chunks": [{"chunk_id": "legacy", "source_type": "hermes_evidence", "text": "核心卖点和核心槽点TOP", "tags": ["核心卖点TOP"]}],
    "compare_rows": [],
    "opportunity_rows": []
}, ensure_ascii=False))
""",
        encoding="utf-8",
    )
    fake_hermes.chmod(0o755)

    result = generate_outputs(
        autohome_input=zj_path,
        dcd_input=dcd_path,
        postprocess_input=tmp_path / "dual.xlsx",
        summary_output=tmp_path / "summary" / "测试车_双平台口碑摘要.xlsx",
        terms_output=tmp_path / "wordcloud" / "测试车_词云词项清单.xlsx",
        wordcloud_output_dir=tmp_path / "wordcloud",
        final_report_output=tmp_path / "ai" / "final_report.json",
        qa_chunks_output=tmp_path / "ai" / "qa_chunks.json",
        model_name="测试车",
        progress_file=tmp_path / "progress" / "generating_hermes_outputs.progress.json",
        summary_script=summary_script,
        wordcloud_script=wordcloud_script,
        hermes_command=str(fake_hermes),
        env={**os.environ, "LLM_API_KEY": "test-key", "LLM_MODEL_REPORT": "deepseek-chat"},
    )

    assert result["status"] == "success"
    report_text = Path(result["final_report_path"]).read_text(encoding="utf-8")
    qa_text = Path(result["qa_chunks_path"]).read_text(encoding="utf-8")
    assert "核心卖点" not in report_text
    assert "核心卖点" not in qa_text
    assert "核心槽点TOP" not in report_text
    assert "核心槽点TOP" not in qa_text
    assert "核心好评" in report_text
    assert "最满意TOP" in report_text
    assert "最不满意TOP" in report_text
