from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from worker_app.postprocess_bridge import prepare_autohome_workbook


def test_prepare_autohome_workbook_bridges_single_sheet_export(tmp_path: Path) -> None:
    source = tmp_path / "autohome.xlsx"
    prepared = tmp_path / "prepared.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "口碑"
    worksheet.append(["数据类型", "来源链接"])
    worksheet.append(["车主购车口碑", "https://example.com/review/1"])
    workbook.save(source)

    effective_path = prepare_autohome_workbook(source, prepared)

    assert effective_path == prepared
    bridged = load_workbook(prepared)
    assert bridged.sheetnames == ["购车口碑"]
    assert bridged["购车口碑"]["A2"].value == "车主购车口碑"
