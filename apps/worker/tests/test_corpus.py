from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from worker_app.corpus import (
    AUTOHOME_HEADERS,
    dedupe_key_for_row,
    export_platform_workbook,
    load_platform_state,
    upsert_platform_rows,
)


def test_dedupe_key_uses_stable_hash_when_source_link_is_missing() -> None:
    row = {
        "用户名": "车主A",
        "发表日期": "2026-05-01",
        "评价详情": "空间够用，能耗满意。",
        "来源链接": "",
    }

    first = dedupe_key_for_row("autohome", row)
    second = dedupe_key_for_row("autohome", dict(row))

    assert first == second
    assert first.startswith("hash:")
    assert dedupe_key_for_row("autohome", {**row, "评价详情": "另一条评论"}) != first


def test_corpus_upserts_new_rows_and_exports_full_platform_workbook(tmp_path: Path) -> None:
    database_url = f"sqlite+pysqlite:///{tmp_path / 'corpus.db'}"
    rows = [
        {
            "用户名": "车主A",
            "发表日期": "2026-05-01",
            "车型": "测试车 2026款",
            "综合口碑": "4.8",
            "评价详情": "第一条评论",
            "来源链接": "https://k.autohome.com.cn/detail/view_01abc.html",
            "抓取页码": "1",
        },
        {
            "用户名": "车主B",
            "发表日期": "2026-05-02",
            "车型": "测试车 2026款",
            "综合口碑": "4.6",
            "评价详情": "第二条评论",
            "来源链接": "",
            "抓取页码": "1",
        },
    ]

    first = upsert_platform_rows(
        database_url=database_url,
        query="测试车",
        model_name="测试车",
        platform="autohome",
        series_id="8089",
        job_id="job_first",
        rows=rows,
    )
    second = upsert_platform_rows(
        database_url=database_url,
        query="测试车",
        model_name="测试车",
        platform="autohome",
        series_id="8089",
        job_id="job_second",
        rows=[rows[0], {**rows[1], "用户名": "车主C", "评价详情": "第三条评论"}],
    )

    assert first.inserted_count == 2
    assert first.total_count == 2
    assert second.inserted_count == 1
    assert second.total_count == 3

    state = load_platform_state(database_url, query="测试车", platform="autohome", series_id="8089")
    assert state.existing_count == 3
    assert state.known_links == {"https://k.autohome.com.cn/detail/view_01abc.html"}

    output = tmp_path / "ZJ测试车原始口碑.xlsx"
    exported_count = export_platform_workbook(
        database_url=database_url,
        query="测试车",
        platform="autohome",
        series_id="8089",
        output_path=output,
        headers=AUTOHOME_HEADERS,
    )

    assert exported_count == 3
    workbook = load_workbook(output)
    sheet = workbook.active
    exported = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(exported) == 3
    header = [cell.value for cell in sheet[1]]
    assert header[:4] == AUTOHOME_HEADERS[:4]
    source_link_index = header.index("来源链接")
    assert exported[0][source_link_index] == "https://k.autohome.com.cn/detail/view_01abc.html"
