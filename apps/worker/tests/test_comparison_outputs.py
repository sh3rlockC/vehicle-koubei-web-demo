from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from worker_app.comparison_outputs import VehicleSnapshot, generate_comparison_outputs


def write_snapshot(tmp_path: Path, model_name: str, facts: list[dict]) -> VehicleSnapshot:
    root = tmp_path / model_name
    root.mkdir(parents=True)
    final_report = root / "final_report.json"
    analysis_facts = root / "analysis_facts.jsonl"
    final_report.write_text(
        json.dumps({"headline": f"{model_name} 总结", "strength_blocks": [], "weakness_blocks": []}, ensure_ascii=False),
        encoding="utf-8",
    )
    analysis_facts.write_text(
        "\n".join(json.dumps(fact, ensure_ascii=False) for fact in facts) + "\n",
        encoding="utf-8",
    )
    return VehicleSnapshot(
        model_name=model_name,
        source_job_id=f"job_{model_name}",
        final_report_path=final_report,
        analysis_facts_path=analysis_facts,
    )


def test_generate_comparison_outputs_writes_dimension_matrix_excel(tmp_path: Path) -> None:
    snapshots = [
        write_snapshot(
            tmp_path,
            "车型A",
            [
                {
                    "comment_id": "a1",
                    "date": "2026-03-01",
                    "section_facts": {"positive": "空间很大，第二排舒服", "negative": "车机卡顿"},
                    "local_keywords": ["空间", "智能化"],
                },
                {
                    "comment_id": "a1",
                    "date": "2026-03-01",
                    "section_facts": {"positive": "后排空间也够用", "negative": ""},
                    "local_keywords": ["空间"],
                },
                {
                    "comment_id": "a2",
                    "date": "2026-03-02",
                    "section_facts": {"positive": "外观好看，内饰做工好，舒适性不错", "negative": "能耗偏高，内饰异味，座椅偏硬"},
                    "local_keywords": ["外观", "能耗", "内饰", "舒适"],
                },
                {
                    "comment_id": "a3",
                    "date": "2026-03-03",
                    "section_facts": {"positive": "座椅舒适，隔音舒适", "negative": "胎噪影响舒适性"},
                    "local_keywords": ["舒适"],
                },
                {
                    "comment_id": "a4",
                    "date": "2026-03-04",
                    "section_facts": {"positive": "", "negative": "舒适性还有颠簸"},
                    "local_keywords": ["舒适"],
                },
            ],
        ),
        write_snapshot(
            tmp_path,
            "车型B",
            [
                {
                    "comment_id": "b1",
                    "date": "2026-03-01",
                    "section_facts": {"positive": "智能驾驶好用，内饰屏幕好，座椅舒适", "negative": "第三排空间小，内饰塑料感，舒适性一般"},
                    "local_keywords": ["智能化", "空间", "内饰", "舒适"],
                },
                {
                    "comment_id": "b2",
                    "date": "2026-03-02",
                    "section_facts": {"positive": "", "negative": "隔音影响舒适"},
                    "local_keywords": ["舒适"],
                }
            ],
        ),
    ]

    result = generate_comparison_outputs(snapshots=snapshots, output_dir=tmp_path / "comparison", env={})

    report = result["report_json"]
    assert report["conclusion"]["summary"]
    assert report["conclusion"]["source"] == "fallback"
    assert [row["dimension"] for row in report["dimensions"]] == [
        "空间",
        "外观",
        "内饰",
        "配置/性价比",
        "续航/能耗",
        "智能化",
        "驾驶感受/操控/动力",
        "舒适性",
    ]
    space = report["dimensions"][0]["vehicles"]
    assert space[0]["positive_mentions"] == 1
    assert space[0]["negative_mentions"] == 0
    assert space[0]["positive_evidence_ids"] == ["a1"]
    assert space[1]["positive_mentions"] == 0
    assert space[1]["negative_mentions"] == 1
    assert report["dimensions"][0]["winner_model_names"] == ["车型A"]
    assert report["dimensions"][0]["winner_score"] == 1.0
    assert report["dimensions"][0]["winner_label"] == "车型A"
    interior = report["dimensions"][2]
    assert interior["winner_model_names"] == ["车型A", "车型B"]
    assert interior["winner_label"] == "车型A、车型B"
    value = report["dimensions"][3]
    assert value["winner_model_names"] == []
    assert value["winner_score"] is None
    assert value["winner_label"] == "无数据"
    comfort = report["dimensions"][7]
    assert comfort["winner_model_names"] == ["车型A"]
    assert comfort["winner_label"] == "车型A"

    dimension_path = tmp_path / "comparison" / "comparison_dimension_matrix.xlsx"
    assert str(dimension_path) in result["artifact_paths"]
    workbook = load_workbook(dimension_path)
    matrix = workbook["维度对比"]
    assert matrix.cell(row=1, column=1).value == "维度"
    assert matrix.cell(row=1, column=2).value == "车型对比胜者"
    assert matrix.cell(row=1, column=3).value == "车型A 优点提及数"
    assert matrix.cell(row=1, column=4).value == "车型A 槽点提及数"
    assert matrix.cell(row=2, column=1).value == "空间"
    assert matrix.cell(row=2, column=2).value == "车型A"
    assert matrix.cell(row=2, column=3).value == 1
    assert matrix.cell(row=2, column=4).value == 0
    assert matrix.cell(row=2, column=3).fill.fgColor.rgb == "FFC6EFCE"
    assert matrix.cell(row=2, column=4).fill.fgColor.rgb == "FFFFC7CE"
    assert matrix.cell(row=9, column=2).value == "车型A"
    conclusion = workbook["结论"]
    assert conclusion.cell(row=1, column=1).value == "LLM 多车型对比结论"
