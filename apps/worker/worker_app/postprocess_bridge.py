from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path


def _copy_sheet(source_sheet, target_sheet) -> None:
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(list(row))


def prepare_autohome_workbook(source_path: Path, prepared_path: Path) -> Path:
    from openpyxl import Workbook, load_workbook

    workbook = load_workbook(source_path)
    if "购车口碑" in workbook.sheetnames or "试驾口碑" in workbook.sheetnames:
        return source_path
    if "口碑" not in workbook.sheetnames:
        raise SystemExit(f"unsupported autohome workbook shape: {source_path}")

    source_sheet = workbook["口碑"]
    prepared_path.parent.mkdir(parents=True, exist_ok=True)
    new_workbook = Workbook()
    target_sheet = new_workbook.active
    target_sheet.title = "购车口碑"
    _copy_sheet(source_sheet, target_sheet)
    new_workbook.save(prepared_path)
    return prepared_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Bridge single-sheet Autohome exports into the koubei-postprocess workbook contract")
    parser.add_argument("--postprocess-script", required=True)
    parser.add_argument("--source-zj-input", required=True)
    parser.add_argument("--prepared-zj-output", required=True)
    parser.add_argument("--dcd-input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    source_zj_input = Path(args.source_zj_input).resolve()
    prepared_zj_output = Path(args.prepared_zj_output).resolve()
    dcd_input = Path(args.dcd_input).resolve()
    output = Path(args.output).resolve()

    effective_zj_input = prepare_autohome_workbook(source_zj_input, prepared_zj_output)
    command = [
        sys.executable,
        args.postprocess_script,
        "--zj-input",
        str(effective_zj_input),
        "--dcd-input",
        str(dcd_input),
        "--output",
        str(output),
    ]
    completed = subprocess.run(command, check=False)
    return completed.returncode


if __name__ == "__main__":
    raise SystemExit(main())
