from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, date, datetime
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    from worker_app.hermes_outputs import _call_aggregate_llm_json
except Exception:  # pragma: no cover - comparison fallback still works without Hermes internals.
    _call_aggregate_llm_json = None


@dataclass(frozen=True)
class VehicleSnapshot:
    model_name: str
    source_job_id: str
    final_report_path: Path
    analysis_facts_path: Path
    llm_metrics_path: Path | None = None


DIMENSION_KEYWORDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("空间", ("空间", "后排", "二排", "第二排", "三排", "第三排", "储物", "后备箱", "行李箱", "乘坐")),
    ("外观", ("外观", "颜值", "造型", "车身", "前脸", "尾灯", "大灯", "颜色")),
    ("内饰", ("内饰", "座舱", "用料", "做工", "中控", "仪表", "屏幕", "氛围灯")),
    ("配置/性价比", ("配置", "性价比", "价格", "权益", "优惠", "标配", "选装", "配置高")),
    ("续航/能耗", ("续航", "能耗", "油耗", "电耗", "耗电", "充电", "补能", "亏电")),
    ("智能化", ("智能", "车机", "智驾", "辅助驾驶", "语音", "导航", "系统", "OTA", "NOA")),
    ("驾驶感受/操控/动力", ("驾驶", "操控", "动力", "加速", "刹车", "底盘", "转向", "悬架", "变速")),
    ("舒适性", ("舒适", "座椅", "隔音", "噪音", "胎噪", "风噪", "震动", "颠簸", "空调")),
)

POSITIVE_FILL = PatternFill("solid", fgColor="FFC6EFCE")
NEGATIVE_FILL = PatternFill("solid", fgColor="FFFFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="FFE2E8F0")


def _read_json(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", str(value or ""))
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _count_jsonl(path: Path, *, start_date: str | None = None, end_date: str | None = None) -> int:
    start = _parse_date(start_date) if start_date else None
    end = _parse_date(end_date) if end_date else None
    count = 0
    try:
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                if start or end:
                    try:
                        payload = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    parsed = _parse_date(payload.get("date") if isinstance(payload, dict) else None)
                    if parsed is None:
                        continue
                    if start and parsed < start:
                        continue
                    if end and parsed > end:
                        continue
                count += 1
    except OSError:
        return 0
    return count


def _iter_facts(path: Path, *, start_date: str | None = None, end_date: str | None = None) -> list[dict[str, Any]]:
    start = _parse_date(start_date) if start_date else None
    end = _parse_date(end_date) if end_date else None
    facts: list[dict[str, Any]] = []
    try:
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                try:
                    payload = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if not isinstance(payload, dict):
                    continue
                parsed = _parse_date(payload.get("date"))
                if start or end:
                    if parsed is None:
                        continue
                    if start and parsed < start:
                        continue
                    if end and parsed > end:
                        continue
                facts.append(payload)
    except OSError:
        return []
    return facts


def _fact_text_for_sentiment(fact: dict[str, Any], sentiment: str) -> str:
    section_facts = fact.get("section_facts") if isinstance(fact.get("section_facts"), dict) else {}
    values = [
        section_facts.get(sentiment),
        fact.get(f"{sentiment}_text"),
    ]
    return " ".join(str(value) for value in values if value)


def _mentioned_dimensions(text: str) -> set[str]:
    normalized = text.lower()
    mentioned: set[str] = set()
    for dimension, keywords in DIMENSION_KEYWORDS:
        if any(keyword.lower() in normalized for keyword in keywords):
            mentioned.add(dimension)
    return mentioned


def _dimension_matrix(
    snapshots: list[VehicleSnapshot],
    *,
    start_date: str | None = None,
    end_date: str | None = None,
) -> list[dict[str, Any]]:
    per_vehicle: dict[str, dict[str, dict[str, set[str]]]] = {}
    for snapshot in snapshots:
        buckets = {
            dimension: {
                "positive_evidence_ids": set(),
                "negative_evidence_ids": set(),
            }
            for dimension, _keywords in DIMENSION_KEYWORDS
        }
        facts = _iter_facts(snapshot.analysis_facts_path, start_date=start_date, end_date=end_date)
        for index, fact in enumerate(facts, start=1):
            comment_id = str(fact.get("comment_id") or f"__row_{index}")
            for dimension in _mentioned_dimensions(_fact_text_for_sentiment(fact, "positive")):
                buckets[dimension]["positive_evidence_ids"].add(comment_id)
            for dimension in _mentioned_dimensions(_fact_text_for_sentiment(fact, "negative")):
                buckets[dimension]["negative_evidence_ids"].add(comment_id)
        per_vehicle[snapshot.model_name] = buckets

    rows: list[dict[str, Any]] = []
    for dimension, _keywords in DIMENSION_KEYWORDS:
        rows.append(
            {
                "dimension": dimension,
                "vehicles": [
                    {
                        "model_name": snapshot.model_name,
                        "positive_mentions": len(
                            per_vehicle[snapshot.model_name][dimension]["positive_evidence_ids"]
                        ),
                        "negative_mentions": len(
                            per_vehicle[snapshot.model_name][dimension]["negative_evidence_ids"]
                        ),
                        "positive_evidence_ids": sorted(
                            per_vehicle[snapshot.model_name][dimension]["positive_evidence_ids"]
                        ),
                        "negative_evidence_ids": sorted(
                            per_vehicle[snapshot.model_name][dimension]["negative_evidence_ids"]
                        ),
                    }
                    for snapshot in snapshots
                ],
            }
        )
    return rows


def _vehicle_summary(snapshot: VehicleSnapshot, *, start_date: str | None = None, end_date: str | None = None) -> dict[str, Any]:
    report = _read_json(snapshot.final_report_path)
    total_count = _count_jsonl(snapshot.analysis_facts_path)
    selected_count = _count_jsonl(snapshot.analysis_facts_path, start_date=start_date, end_date=end_date) if start_date or end_date else total_count
    return {
        "model_name": snapshot.model_name,
        "source_job_id": snapshot.source_job_id,
        "comment_fact_count": selected_count,
        "total_comment_fact_count": total_count,
        "headline": report.get("headline") or report.get("summary") or f"{snapshot.model_name} 口碑摘要",
        "structured_sections": report.get("structured_sections") or report.get("sections") or {},
        "source_report_path": str(snapshot.final_report_path),
        "source_facts_path": str(snapshot.analysis_facts_path),
    }


def _write_summary_workbook(path: Path, vehicles: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "竞品对比"
    sheet.append(["车型", "评论事实数", "摘要"])
    for vehicle in vehicles:
        sheet.append([vehicle["model_name"], vehicle["comment_fact_count"], str(vehicle["headline"])])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _fallback_conclusion(vehicles: list[dict[str, Any]], dimensions: list[dict[str, Any]]) -> dict[str, Any]:
    leaders: list[str] = []
    risk_points: list[str] = []
    for row in dimensions:
        dimension = row["dimension"]
        vehicle_rows = row["vehicles"]
        if not vehicle_rows:
            continue
        positive_leader = max(vehicle_rows, key=lambda item: item["positive_mentions"])
        negative_leader = max(vehicle_rows, key=lambda item: item["negative_mentions"])
        if positive_leader["positive_mentions"] > 0:
            leaders.append(f"{dimension}上{positive_leader['model_name']}优点提及最多")
        if negative_leader["negative_mentions"] > 0:
            risk_points.append(f"{dimension}上{negative_leader['model_name']}槽点提及最多")

    vehicle_names = "、".join(str(vehicle["model_name"]) for vehicle in vehicles)
    summary = f"{vehicle_names}的口碑差异主要体现在固定维度的优劣提及分布上。"
    if leaders:
        summary += leaders[0] + "。"
    rationale = (leaders[:3] + risk_points[:3])[:5] or ["各车型在固定维度上的提及量接近，需要结合原始口碑继续查看。"]
    return {"summary": summary, "rationale_bullets": rationale, "source": "fallback"}


def _build_conclusion_prompt(vehicles: list[dict[str, Any]], dimensions: list[dict[str, Any]]) -> str:
    payload = {"vehicles": vehicles, "dimensions": dimensions}
    return (
        "你是汽车口碑竞品分析Agent。只根据输入JSON总结多车型对比结论，不要引入外部资料。"
        "输出严格JSON：{\"summary\":\"...\",\"rationale_bullets\":[\"...\"]}。"
        "summary限制在120字内，rationale_bullets给3-5条。\n"
        + json.dumps(payload, ensure_ascii=False, default=str)
    )


def _llm_enabled(env: dict[str, str]) -> bool:
    if str(env.get("COMPARISON_LLM_ENABLED", "1")).lower() in {"0", "false", "no"}:
        return False
    return bool(env.get("LLM_API_KEY") and (env.get("LLM_MODEL_REPORT") or env.get("LLM_MODEL_BATCH")))


def _comparison_conclusion(
    *,
    vehicles: list[dict[str, Any]],
    dimensions: list[dict[str, Any]],
    output_dir: Path,
    env: dict[str, str],
) -> tuple[dict[str, Any], str | None]:
    fallback = _fallback_conclusion(vehicles, dimensions)
    if not _llm_enabled(env) or _call_aggregate_llm_json is None:
        return fallback, None

    prompt = _build_conclusion_prompt(vehicles, dimensions)
    try:
        payload = _call_aggregate_llm_json(
            prompt,
            hermes_command=env.get("HERMES_COMMAND", "hermes"),
            env=env,
            model=env.get("LLM_MODEL_REPORT") or env.get("LLM_MODEL_BATCH") or "",
            debug_dir=output_dir / "hermes_debug",
            call_label="comparison_conclusion",
            metrics=None,
            metric_stage="comparison_conclusion",
        )
        if not isinstance(payload, dict):
            raise ValueError("comparison_conclusion payload is not object")
        summary = str(payload.get("summary") or "").strip()
        bullets = payload.get("rationale_bullets")
        if not summary or not isinstance(bullets, list):
            raise ValueError("comparison_conclusion missing summary or rationale_bullets")
        return {
            "summary": summary,
            "rationale_bullets": [str(item) for item in bullets if str(item).strip()][:5],
            "source": "llm",
        }, None
    except Exception as exc:
        fallback["fallback_reason"] = str(exc)
        return fallback, str(exc)


def _write_dimension_workbook(path: Path, *, conclusion: dict[str, Any], dimensions: list[dict[str, Any]], vehicles: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    matrix = workbook.active
    matrix.title = "维度对比"
    header = ["维度"]
    for vehicle in vehicles:
        model_name = str(vehicle["model_name"])
        header.extend([f"{model_name} 优点提及数", f"{model_name} 槽点提及数"])
    matrix.append(header)
    for cell in matrix[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row in dimensions:
        values: list[Any] = [row["dimension"]]
        for vehicle in row["vehicles"]:
            values.extend([vehicle["positive_mentions"], vehicle["negative_mentions"]])
        matrix.append(values)
        row_index = matrix.max_row
        for column_index in range(2, len(values) + 1):
            matrix.cell(row=row_index, column=column_index).fill = POSITIVE_FILL if column_index % 2 == 0 else NEGATIVE_FILL

    matrix.freeze_panes = "B2"
    matrix.column_dimensions["A"].width = 22
    for column_cells in matrix.iter_cols(min_col=2, max_col=matrix.max_column):
        matrix.column_dimensions[column_cells[0].column_letter].width = 18

    conclusion_sheet = workbook.create_sheet("结论")
    conclusion_sheet.append(["LLM 多车型对比结论"])
    conclusion_sheet.append([conclusion.get("summary", "")])
    conclusion_sheet.append(["依据"])
    for item in conclusion.get("rationale_bullets") or []:
        conclusion_sheet.append([item])
    conclusion_sheet["A1"].font = Font(bold=True)
    conclusion_sheet.column_dimensions["A"].width = 90

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def generate_comparison_outputs(
    *,
    snapshots: list[VehicleSnapshot],
    output_dir: Path,
    start_date: str | None = None,
    end_date: str | None = None,
    env: dict[str, str] | None = None,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    active_env = env or {}
    vehicles = [_vehicle_summary(snapshot, start_date=start_date, end_date=end_date) for snapshot in snapshots]
    dimensions = _dimension_matrix(snapshots, start_date=start_date, end_date=end_date)
    conclusion, conclusion_fallback_reason = _comparison_conclusion(
        vehicles=vehicles,
        dimensions=dimensions,
        output_dir=output_dir,
        env=active_env,
    )
    generated_at = datetime.now(UTC).isoformat()
    date_range = {"start_date": start_date, "end_date": end_date} if start_date or end_date else None
    report_json = {
        "headline": "竞品口碑对比",
        "generated_at": generated_at,
        "source": "hermes_comparison",
        "date_range": date_range,
        "vehicle_count": len(vehicles),
        "vehicles": vehicles,
        "dimensions": dimensions,
        "conclusion": conclusion,
        "comparison": {
            "summary": conclusion["summary"],
            "dimensions": [dimension for dimension, _keywords in DIMENSION_KEYWORDS],
        },
    }

    final_json = output_dir / "final_comparison.json"
    summary_xlsx = output_dir / "comparison_summary.xlsx"
    dimension_xlsx = output_dir / "comparison_dimension_matrix.xlsx"
    metrics_json = output_dir / "llm_metrics.json"
    final_json.write_text(json.dumps(report_json, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_summary_workbook(summary_xlsx, vehicles)
    _write_dimension_workbook(dimension_xlsx, conclusion=conclusion, dimensions=dimensions, vehicles=vehicles)
    metrics_json.write_text(
        json.dumps(
            {
                "source": "hermes_comparison",
                "generated_at": generated_at,
                "provider": active_env.get("LLM_PROVIDER"),
                "model_report": active_env.get("LLM_MODEL_REPORT"),
                "vehicle_count": len(vehicles),
                "conclusion_source": conclusion["source"],
                "conclusion_fallback_reason": conclusion_fallback_reason,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return {
        "report_json": report_json,
        "artifact_paths": [str(final_json), str(summary_xlsx), str(dimension_xlsx), str(metrics_json)],
        "degraded": bool(conclusion_fallback_reason),
    }
