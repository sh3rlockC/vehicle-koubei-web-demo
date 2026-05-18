from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
import hashlib
import json
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import JSON as SAJSON
from sqlalchemy import Column, DateTime, Integer, MetaData, String, Table, Text, UniqueConstraint, bindparam, create_engine, func, select

AUTOHOME_HEADERS = [
    "数据类型",
    "用户名",
    "发表日期",
    "口碑标题",
    "综合口碑",
    "车型",
    "行驶里程",
    "电耗",
    "裸车购买价",
    "参考价格",
    "购买时间",
    "探店时间",
    "购买地点",
    "探店地点",
    "评价详情",
    "来源链接",
    "抓取页码",
]

DCD_HEADERS = [
    "用户名",
    "用户标签",
    "评价车型",
    "懂车分",
    "发布时间",
    "用户评分",
    "续航",
    "购车时间",
    "裸车价",
    "购车地",
    "评价全文",
    "来源链接",
    "抓取页码",
]

PLATFORM_HEADERS = {
    "autohome": AUTOHOME_HEADERS,
    "dongchedi": DCD_HEADERS,
}

INCREMENTAL_MAX_SCAN_PAGES = 10
INCREMENTAL_STOP_AFTER_KNOWN_PAGES = 2

metadata = MetaData()
koubei_raw_comments = Table(
    "koubei_raw_comments",
    metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("query_key", String(255), nullable=False),
    Column("query", String(255), nullable=False),
    Column("model_name", String(255), nullable=False),
    Column("platform", String(32), nullable=False),
    Column("series_id", String(64), nullable=False),
    Column("source_link", Text, nullable=True),
    Column("dedupe_key", String(128), nullable=False),
    Column("row_json", SAJSON, nullable=False),
    Column("first_seen_job_id", String(64), nullable=True),
    Column("last_seen_job_id", String(64), nullable=True),
    Column("first_seen_at", DateTime(timezone=True), nullable=False),
    Column("last_seen_at", DateTime(timezone=True), nullable=False),
    Column("published_at", String(32), nullable=True),
    Column("page", Integer, nullable=True),
    UniqueConstraint("query_key", "platform", "series_id", "dedupe_key", name="uq_koubei_raw_comment_dedupe"),
)


@dataclass(frozen=True)
class PlatformCorpusState:
    existing_count: int
    known_links: set[str]


@dataclass(frozen=True)
class CorpusImportResult:
    inserted_count: int
    updated_count: int
    total_count: int


def _engine_kwargs(database_url: str) -> dict[str, Any]:
    if database_url.startswith("sqlite"):
        return {"connect_args": {"check_same_thread": False}}
    return {}


def create_corpus_engine(database_url: str):
    return create_engine(database_url, future=True, **_engine_kwargs(database_url))


def ensure_corpus_schema(engine) -> None:
    metadata.create_all(engine, tables=[koubei_raw_comments])


def query_key(query: str) -> str:
    return re.sub(r"\s+", " ", query).strip().lower()


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _source_link(row: dict[str, Any]) -> str:
    return _clean(row.get("来源链接"))


def dedupe_key_for_row(platform: str, row: dict[str, Any]) -> str:
    source_link = _source_link(row)
    if source_link:
        return f"link:{source_link}"

    username = _clean(row.get("用户名"))
    published_at = _clean(row.get("发表日期") or row.get("发布时间"))
    body = _clean(row.get("评价详情") or row.get("评价全文"))
    digest = hashlib.sha256(f"{platform}|{username}|{published_at}|{body}".encode("utf-8")).hexdigest()
    return f"hash:{digest}"


def _published_at(row: dict[str, Any]) -> str | None:
    value = _clean(row.get("发表日期") or row.get("发布时间"))
    return value or None


def _page(row: dict[str, Any]) -> int | None:
    raw = row.get("抓取页码")
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _row_json(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def load_platform_state(database_url: str, *, query: str, platform: str, series_id: str) -> PlatformCorpusState:
    engine = create_corpus_engine(database_url)
    ensure_corpus_schema(engine)
    key = query_key(query)
    with engine.begin() as conn:
        existing_count = conn.execute(
            select(func.count())
            .select_from(koubei_raw_comments)
            .where(
                koubei_raw_comments.c.query_key == key,
                koubei_raw_comments.c.platform == platform,
                koubei_raw_comments.c.series_id == str(series_id),
            )
        ).scalar_one()
        rows = conn.execute(
            select(koubei_raw_comments.c.source_link).where(
                koubei_raw_comments.c.query_key == key,
                koubei_raw_comments.c.platform == platform,
                koubei_raw_comments.c.series_id == str(series_id),
                koubei_raw_comments.c.source_link.is_not(None),
            )
        ).all()
    return PlatformCorpusState(
        existing_count=int(existing_count or 0),
        known_links={str(row[0]).strip() for row in rows if str(row[0] or "").strip()},
    )


def upsert_platform_rows(
    *,
    database_url: str,
    query: str,
    model_name: str,
    platform: str,
    series_id: str,
    job_id: str,
    rows: list[dict[str, Any]],
) -> CorpusImportResult:
    engine = create_corpus_engine(database_url)
    ensure_corpus_schema(engine)
    key = query_key(query)
    now = datetime.now(UTC)
    inserted = 0
    updated = 0
    insert_statement = koubei_raw_comments.insert()
    update_statement = (
        koubei_raw_comments.update()
        .where(koubei_raw_comments.c.id == bindparam("comment_id"))
        .values(
            query=query.strip(),
            model_name=model_name,
            row_json=bindparam("row_json"),
            last_seen_job_id=job_id,
            last_seen_at=now,
            published_at=bindparam("row_published_at"),
            page=bindparam("row_page"),
        )
    )

    with engine.begin() as conn:
        for raw_row in rows:
            row = {str(key): "" if value is None else value for key, value in raw_row.items()}
            dedupe_key = dedupe_key_for_row(platform, row)
            existing_id = conn.execute(
                select(koubei_raw_comments.c.id).where(
                    koubei_raw_comments.c.query_key == key,
                    koubei_raw_comments.c.platform == platform,
                    koubei_raw_comments.c.series_id == str(series_id),
                    koubei_raw_comments.c.dedupe_key == dedupe_key,
                )
            ).scalar_one_or_none()
            payload = {
                "query_key": key,
                "query": query.strip(),
                "model_name": model_name,
                "platform": platform,
                "series_id": str(series_id),
                "source_link": _source_link(row) or None,
                "dedupe_key": dedupe_key,
                "row_json": row,
                "first_seen_job_id": job_id,
                "last_seen_job_id": job_id,
                "first_seen_at": now,
                "last_seen_at": now,
                "published_at": _published_at(row),
                "page": _page(row),
            }
            if existing_id is None:
                conn.execute(insert_statement, payload)
                inserted += 1
            else:
                conn.execute(
                    update_statement,
                    {
                        "comment_id": existing_id,
                        "row_json": row,
                        "row_published_at": _published_at(row),
                        "row_page": _page(row),
                    },
                )
                updated += 1

        total_count = conn.execute(
            select(func.count())
            .select_from(koubei_raw_comments)
            .where(
                koubei_raw_comments.c.query_key == key,
                koubei_raw_comments.c.platform == platform,
                koubei_raw_comments.c.series_id == str(series_id),
            )
        ).scalar_one()

    return CorpusImportResult(inserted_count=inserted, updated_count=updated, total_count=int(total_count or 0))


def read_workbook_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    workbook = load_workbook(path)
    sheet = workbook["口碑明细"] if "口碑明细" in workbook.sheetnames else workbook.active
    headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: "" if value is None else value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if any(str(value).strip() for value in row.values()):
            rows.append(row)
    return rows


def _apply_basic_sheet_style(sheet, platform: str) -> None:
    for cell in sheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    if platform == "autohome":
        sheet.column_dimensions["O"].width = 110
        sheet.column_dimensions["P"].width = 60
    else:
        sheet.column_dimensions["K"].width = 100
        sheet.column_dimensions["L"].width = 60


def export_platform_workbook(
    *,
    database_url: str,
    query: str,
    platform: str,
    series_id: str,
    output_path: Path,
    headers: list[str],
) -> int:
    engine = create_corpus_engine(database_url)
    ensure_corpus_schema(engine)
    key = query_key(query)
    with engine.begin() as conn:
        rows = conn.execute(
            select(koubei_raw_comments.c.row_json)
            .where(
                koubei_raw_comments.c.query_key == key,
                koubei_raw_comments.c.platform == platform,
                koubei_raw_comments.c.series_id == str(series_id),
            )
            .order_by(koubei_raw_comments.c.id.asc())
        ).all()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "口碑明细" if platform == "dongchedi" else "口碑"
    sheet.append(headers)
    for row in rows:
        item = _row_json(row[0])
        sheet.append([item.get(header, "") for header in headers])
    _apply_basic_sheet_style(sheet, platform)
    workbook.save(output_path)
    return len(rows)


def write_known_links_file(path: Path, links: set[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(sorted(links)) + ("\n" if links else ""), encoding="utf-8")


def read_validation_incremental_stats(validation_path: Path) -> dict[str, Any]:
    if not validation_path.exists():
        return {}
    try:
        payload = json.loads(validation_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    incremental = payload.get("incremental")
    if isinstance(incremental, dict):
        return incremental
    page_counts = payload.get("page_counts") or payload.get("page_link_counts") or {}
    return {"pages_scanned": len(page_counts), "stop_reason": None}
