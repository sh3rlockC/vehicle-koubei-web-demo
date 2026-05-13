from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import signal
import subprocess
import sys
import time

if __package__ in {None, ""}:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if sys.path and os.path.abspath(sys.path[0]) == script_dir:
        sys.path.pop(0)
    parent_dir = os.path.dirname(script_dir)
    if parent_dir not in sys.path:
        sys.path.insert(0, parent_dir)

from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from threading import Lock
from threading import current_thread, main_thread
from typing import Any

import requests
from openpyxl import Workbook, load_workbook


PLATFORM_AUTOHOME = "汽车之家"
PLATFORM_DCD = "懂车帝"
DEFAULT_BATCH_TARGET_BYTES = 45_000
DEFAULT_BATCH_CONCURRENCY = 3
DEFAULT_BATCH_MODEL = "deepseek-v4-flash"
DEFAULT_REPORT_MODEL = "deepseek-v4-pro"
MAX_FACT_TEXT_CHARS = 240
MAX_FACT_SENTIMENT_CHARS = 70
MAX_FACT_SECTION_CHARS = 35
MAX_FACT_TOPIC_SECTIONS = 2
MAX_REPAIR_RESPONSE_CHARS = 120_000
MAX_AGGREGATE_THEMES_PER_DIRECTION = 2
MAX_AGGREGATE_SUGGESTIONS = 1
MAX_AGGREGATE_PLATFORM_NOTES = 1
MAX_AGGREGATE_PAYLOAD_BYTES = 70_000
DEEPSEEK_RETRYABLE_STATUS_CODES = {429, 500, 502, 503, 504}
DEEPSEEK_MODEL_ALIASES = {
    "deepseekv4pro": "deepseek-v4-pro",
    "deepseek-v4pro": "deepseek-v4-pro",
    "deepseek_v4_pro": "deepseek-v4-pro",
    "deepseekv4flash": "deepseek-v4-flash",
    "deepseek-v4flash": "deepseek-v4-flash",
    "deepseek_v4_flash": "deepseek-v4-flash",
}
LEGACY_LABEL_REPLACEMENTS = {
    "核心卖点TOP": "最满意TOP",
    "核心槽点TOP": "最不满意TOP",
    "核心卖点": "核心好评",
}
COMMENT_SECTION_ALIASES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("positive", ("最满意", "满意", "优点", "优势", "正向反馈")),
    ("negative", ("最不满意", "不满意", "缺点", "槽点", "负向反馈")),
    ("space", ("空间", "乘坐空间", "储物空间")),
    ("driving", ("驾驶感受", "驾驶体验", "操控", "动力", "底盘")),
    ("range", ("续航", "能耗", "油耗", "电耗", "补能")),
    ("appearance", ("外观", "造型")),
    ("interior", ("内饰", "座舱")),
    ("comfort", ("舒适性", "舒适")),
    ("configuration", ("配置", "功能")),
    ("intelligence", ("智能化", "车机", "智驾", "智能驾驶", "辅助驾驶")),
    ("cost_value", ("性价比", "价格", "购车价")),
    ("safety", ("安全", "安全性")),
    ("service", ("服务", "售后", "交付")),
)
COMMENT_SECTION_LABEL_TO_KEY = {
    label: key
    for key, labels in COMMENT_SECTION_ALIASES
    for label in labels
}
COMMENT_SECTION_LABEL_PATTERN = "|".join(
    re.escape(label)
    for label in sorted(COMMENT_SECTION_LABEL_TO_KEY, key=len, reverse=True)
)
COMMENT_SECTION_RE = re.compile(
    rf"(?:【(?P<bracket>{COMMENT_SECTION_LABEL_PATTERN})】|\[(?P<square>{COMMENT_SECTION_LABEL_PATTERN})\]|(?P<plain>{COMMENT_SECTION_LABEL_PATTERN})\s*[：:])"
)
LOCAL_KEYWORD_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("空间", ("空间", "二排", "三排", "后排", "座椅", "储物", "宽敞")),
    ("动力", ("动力", "加速", "提速", "发动机", "电机", "爬坡")),
    ("续航能耗", ("续航", "能耗", "油耗", "电耗", "馈电", "充电")),
    ("智能化", ("车机", "智能", "智驾", "辅助驾驶", "语音", "导航", "NOA")),
    ("舒适性", ("舒适", "底盘", "悬架", "胎噪", "风噪", "隔音", "震动", "NVH")),
    ("外观内饰", ("外观", "造型", "内饰", "用料", "座舱", "屏幕")),
    ("配置价格", ("配置", "价格", "性价比", "优惠", "权益", "置换")),
    ("服务交付", ("售后", "服务", "交付", "销售", "门店", "维修")),
    ("安全", ("安全", "刹车", "制动", "碰撞", "气囊")),
)
SECTION_KEY_TO_LABEL = {
    "positive": "最满意",
    "negative": "最不满意",
    "space": "空间",
    "driving": "驾驶",
    "range": "续航能耗",
    "appearance": "外观",
    "interior": "内饰",
    "comfort": "舒适性",
    "configuration": "配置",
    "intelligence": "智能化",
    "cost_value": "性价比",
    "safety": "安全",
    "service": "服务",
}
LLM_METRICS_LOCK = Lock()


class _AggregateHardTimeout(BaseException):
    pass


@dataclass(frozen=True)
class NormalizedHermesResult:
    report: dict[str, Any]
    keyword_rankings: dict[str, list[dict[str, Any]]]
    qa_chunks: list[dict[str, Any]]
    one_pager_lines: list[str]
    compare_rows: list[dict[str, str]]
    opportunity_rows: list[dict[str, str]]


def _clean_text(value: object, *, limit: int = 900) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def _replace_legacy_labels(value: Any) -> Any:
    if isinstance(value, str):
        text = value
        for old, new in LEGACY_LABEL_REPLACEMENTS.items():
            text = text.replace(old, new)
        return text
    if isinstance(value, list):
        return [_replace_legacy_labels(item) for item in value]
    if isinstance(value, dict):
        return {
            _replace_legacy_labels(key) if isinstance(key, str) else key: _replace_legacy_labels(item)
            for key, item in value.items()
        }
    return value


def _clean_generated_text(value: object, *, limit: int = 900) -> str:
    return _replace_legacy_labels(_clean_text(value, limit=limit))


def _first_value(row: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = _clean_text(row.get(key))
        if value:
            return value
    return ""


def _iter_sheet_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            raw_rows = list(worksheet.iter_rows(values_only=True))
            if not raw_rows:
                continue
            header = [_clean_text(cell) for cell in raw_rows[0]]
            if not any(header):
                continue
            for raw_row in raw_rows[1:]:
                row = {header[index]: raw_row[index] for index in range(min(len(header), len(raw_row))) if header[index]}
                if any(_clean_text(value) for value in row.values()):
                    rows.append(row)
    finally:
        workbook.close()
    return rows


def _append_section(sections: dict[str, str], key: str, value: str) -> None:
    text = _clean_text(value, limit=1600)
    if not text:
        return
    if sections.get(key):
        sections[key] = _clean_text(f"{sections[key]}；{text}", limit=1600)
    else:
        sections[key] = text


def _extract_comment_sections(full_text: str) -> dict[str, str]:
    text = _clean_text(full_text, limit=6000)
    if not text:
        return {}
    matches = list(COMMENT_SECTION_RE.finditer(text))
    sections: dict[str, str] = {}
    for index, match in enumerate(matches):
        label = match.group("bracket") or match.group("square") or match.group("plain")
        key = COMMENT_SECTION_LABEL_TO_KEY.get(label or "")
        if not key:
            continue
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        _append_section(sections, key, text[start:end])
    return sections


def _standard_raw_comment(*, positive_text: str, negative_text: str, full_text: str) -> dict[str, Any]:
    sections = _extract_comment_sections(full_text)
    effective_positive = positive_text or sections.get("positive", "")
    effective_negative = negative_text or sections.get("negative", "")
    topic_sections = {
        key: value
        for key, value in sections.items()
        if key not in {"positive", "negative"} and value
    }
    return {
        "positive_text": effective_positive,
        "negative_text": effective_negative,
        "full_text": full_text,
        "sections": topic_sections,
    }


def _comment_from_row(row: dict[str, Any], *, platform: str, model_name: str) -> dict[str, Any] | None:
    positive_text = _first_value(row, ("最满意", "满意", "优点", "优势", "正向反馈"))
    negative_text = _first_value(row, ("最不满意", "不满意", "缺点", "槽点", "负向反馈"))
    full_text = _first_value(row, ("评价详情", "评价全文", "口碑内容", "内容", "正文", "评论", "原文"))
    date = _first_value(row, ("发表日期", "发布时间", "日期", "时间"))
    row_model = _first_value(row, ("车型", "评价车型", "车款", "车系", "车型名称")) or model_name

    if not any([positive_text, negative_text, full_text]):
        return None
    raw_comment = _standard_raw_comment(positive_text=positive_text, negative_text=negative_text, full_text=full_text)
    return {
        "platform": platform,
        "date": date,
        "model_name": row_model,
        "positive_text": raw_comment["positive_text"],
        "negative_text": raw_comment["negative_text"],
        "full_text": full_text,
        "raw_comment": raw_comment,
    }


def extract_whitelisted_comments(
    *,
    autohome_input: str | Path,
    dcd_input: str | Path | None,
    model_name: str,
) -> list[dict[str, Any]]:
    comments: list[dict[str, Any]] = []
    autohome_path = Path(autohome_input)
    if autohome_path.exists():
        autohome_index = 1
        for row in _iter_sheet_rows(autohome_path):
            comment = _comment_from_row(row, platform=PLATFORM_AUTOHOME, model_name=model_name)
            if comment:
                comment["comment_id"] = f"autohome_{autohome_index:04d}"
                comments.append(comment)
                autohome_index += 1

    dcd_path = Path(dcd_input) if dcd_input else None
    if dcd_path and dcd_path.exists():
        dcd_index = 1
        for row in _iter_sheet_rows(dcd_path):
            comment = _comment_from_row(row, platform=PLATFORM_DCD, model_name=model_name)
            if comment:
                comment["comment_id"] = f"dcd_{dcd_index:04d}"
                comments.append(comment)
                dcd_index += 1

    return comments


def _parse_comment_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value, limit=80)
    if not text:
        return None
    match = re.search(r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _parse_date_param(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def _filter_comments_for_date_range(comments: list[dict[str, str]], *, start_date: str, end_date: str) -> list[dict[str, str]]:
    start = _parse_date_param(start_date)
    end = _parse_date_param(end_date)
    if start > end:
        return []
    selected: list[tuple[date, dict[str, str]]] = []
    for comment in comments:
        parsed = _parse_comment_date(comment.get("date"))
        if parsed is not None and start <= parsed <= end:
            selected.append((parsed, comment))
    return [comment for _parsed, comment in sorted(selected, key=lambda item: (item[0], item[1].get("platform", ""), item[1].get("comment_id", "")))]


def _write_progress(progress_file: Path, *, percent: int, message: str, degraded: bool = False) -> None:
    progress_file.parent.mkdir(parents=True, exist_ok=True)
    progress_file.write_text(
        json.dumps(
            {
                "percent": max(0, min(100, percent)),
                "message": message,
                "degraded": degraded,
                "updated_at": int(time.time()),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def _json_default(value: object) -> str:
    return str(value)


def _strip_hermes_stdout_noise(text: str) -> str:
    lines = []
    for line in text.strip().splitlines():
        stripped = line.strip()
        if stripped.startswith("⚠️ Normalized model ") or stripped.startswith("Normalized model "):
            continue
        lines.append(line)
    return "\n".join(lines).strip()


def _escape_unescaped_inner_quotes(text: str) -> str:
    result: list[str] = []
    in_string = False
    escaped = False
    for index, char in enumerate(text):
        if not in_string:
            result.append(char)
            if char == '"':
                in_string = True
            continue

        if escaped:
            result.append(char)
            escaped = False
            continue
        if char == "\\":
            result.append(char)
            escaped = True
            continue
        if char != '"':
            result.append(char)
            continue

        lookahead = index + 1
        while lookahead < len(text) and text[lookahead] in " \t\r\n":
            lookahead += 1
        if lookahead >= len(text) or text[lookahead] in {":", ",", "}", "]"}:
            result.append(char)
            in_string = False
        else:
            result.append('\\"')
    return "".join(result)


def _json_candidate_slices(text: str) -> list[str]:
    stripped = _strip_hermes_stdout_noise(text)
    candidates = [stripped]
    for fenced in re.finditer(r"```(?:json)?\s*(.*?)```", stripped, flags=re.DOTALL | re.IGNORECASE):
        candidates.append(_strip_hermes_stdout_noise(fenced.group(1)))

    start_positions = [position for position in (stripped.find("{"), stripped.find("[")) if position >= 0]
    if start_positions:
        start = min(start_positions)
        opener = stripped[start]
        closer = "}" if opener == "{" else "]"
        end = stripped.rfind(closer)
        if end > start:
            candidates.append(stripped[start : end + 1])

    unique_candidates: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        cleaned = candidate.strip()
        if cleaned and cleaned not in seen:
            unique_candidates.append(cleaned)
            seen.add(cleaned)
    return unique_candidates


def _json_loads_lenient(candidate: str) -> Any:
    try:
        return json.loads(candidate)
    except json.JSONDecodeError:
        repaired = _escape_unescaped_inner_quotes(candidate)
        if repaired != candidate:
            return json.loads(repaired)
        raise


def _extract_json(text: str) -> Any:
    stripped = _strip_hermes_stdout_noise(text)
    if not stripped:
        raise ValueError("empty response")

    parse_error: Exception | None = None
    for candidate in _json_candidate_slices(stripped):
        try:
            return _json_loads_lenient(candidate)
        except Exception as exc:
            parse_error = exc

    start_positions = [position for position in (stripped.find("{"), stripped.find("[")) if position >= 0]
    if not start_positions:
        raise ValueError("response did not contain JSON")
    start = min(start_positions)
    opener = stripped[start]
    closer = "}" if opener == "{" else "]"
    depth = 0
    in_string = False
    escaped = False
    for index in range(start, len(stripped)):
        char = stripped[index]
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
            continue
        if char == '"':
            in_string = True
        elif char == opener:
            depth += 1
        elif char == closer:
            depth -= 1
            if depth == 0:
                try:
                    return _json_loads_lenient(stripped[start : index + 1])
                except Exception as exc:
                    parse_error = exc
                    break
    raise ValueError("response JSON was not balanced")


def _normalize_deepseek_model(model: str | None) -> str:
    text = (model or "").strip()
    if not text:
        return ""
    return DEEPSEEK_MODEL_ALIASES.get(text.lower().replace(" ", ""), text)


def _runtime_models(env: dict[str, str]) -> tuple[str, str]:
    batch_model = _normalize_deepseek_model(env.get("LLM_MODEL_BATCH") or DEFAULT_BATCH_MODEL)
    report_model = _normalize_deepseek_model(env.get("LLM_MODEL_REPORT") or env.get("LLM_MODEL_QA") or DEFAULT_REPORT_MODEL)
    return batch_model or DEFAULT_BATCH_MODEL, report_model or DEFAULT_REPORT_MODEL


def _runtime_provider(env: dict[str, str]) -> tuple[str, str, dict[str, str]]:
    provider = (env.get("LLM_PROVIDER") or "deepseek").strip().lower()
    model = _normalize_deepseek_model(env.get("LLM_MODEL_REPORT") or env.get("LLM_MODEL_QA") or DEFAULT_REPORT_MODEL)
    base_url = (env.get("LLM_BASE_URL") or "").strip()
    api_key = (env.get("LLM_API_KEY") or "").strip()
    updated_env = dict(env)

    if provider in {"deepseek", "deepseek-v4", "deepseekv4", "deepseekv4pro"}:
        updated_env.setdefault("DEEPSEEK_API_KEY", api_key)
        return "deepseek", model, updated_env

    if base_url:
        return "custom:vehicle-koubei", model, updated_env
    return provider or "deepseek", model, updated_env


def _hermes_llm_mode(env: dict[str, str]) -> str:
    return (env.get("HERMES_LLM_MODE") or "api").strip().lower()


def _deepseek_base_url(env: dict[str, str]) -> str:
    return (env.get("LLM_BASE_URL") or "https://api.deepseek.com").strip().rstrip("/")


def _int_env(env: dict[str, str], key: str, default: int, *, minimum: int = 1) -> int:
    try:
        return max(int(env.get(key) or default), minimum)
    except (TypeError, ValueError):
        return max(default, minimum)


def _write_hermes_config(env: dict[str, str]) -> dict[str, str]:
    provider, model, updated_env = _runtime_provider(env)
    home = Path(updated_env.get("HOME") or str(Path.home())).expanduser()
    hermes_home = home / ".hermes"
    hermes_home.mkdir(parents=True, exist_ok=True)
    base_url = (updated_env.get("LLM_BASE_URL") or "").strip()

    lines = [
        "model:",
        f'  default: "{model}"',
        f'  provider: "{provider}"',
    ]
    if base_url:
        lines.extend(
            [
                "custom_providers:",
                "  - name: vehicle-koubei",
                f'    base_url: "{base_url.rstrip("/")}"',
                "    key_env: LLM_API_KEY",
                "    api_mode: chat_completions",
            ]
        )
    (hermes_home / "config.yaml").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return updated_env


def _hermes_command_path(command: str) -> str | None:
    if "/" in command:
        return command if Path(command).exists() else None
    return shutil.which(command)


def _safe_debug_label(label: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", label).strip("._")
    return safe or "hermes"


def _write_hermes_debug(debug_dir: Path | None, *, label: str, attempt: int, kind: str, content: str) -> None:
    if debug_dir is None:
        return
    debug_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{_safe_debug_label(label)}.attempt_{attempt}.{kind}.txt"
    (debug_dir / filename).write_text(content, encoding="utf-8")


def _reset_hermes_debug_dir(debug_dir: Path) -> None:
    if not debug_dir.exists():
        return
    for path in debug_dir.glob("*.txt"):
        if path.is_file():
            path.unlink()


def _process_output_text(value: str | bytes | None) -> str:
    if value is None:
        return ""
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return value


def _build_json_repair_prompt(*, raw_response: str, parse_error: Exception) -> str:
    response = raw_response.strip()
    if len(response) > MAX_REPAIR_RESPONSE_CHARS:
        response = response[:MAX_REPAIR_RESPONSE_CHARS]
    return (
        "你是 JSON 语法修复器。下面是上一轮模型输出，json.loads 解析失败。\n"
        "任务：只修复 JSON 语法错误，保留原有字段、层级、文本含义和证据 id；不要新增事实，不要重新分析，"
        "不要输出 Markdown、代码块或解释文字。字符串内部不要使用英文双引号；如需引用短语，请改用中文引号「」。\n"
        "允许修复的问题包括：缺失逗号、缺失引号、括号不配平、尾随逗号、非法控制字符。\n"
        f"解析错误：{parse_error}\n"
        "上一轮输出如下：\n"
        f"{response}"
    )


def _comment_for_llm(comment: dict[str, Any]) -> dict[str, Any]:
    raw_comment = comment.get("raw_comment")
    if not isinstance(raw_comment, dict):
        raw_comment = _standard_raw_comment(
            positive_text=_clean_text(comment.get("positive_text"), limit=900),
            negative_text=_clean_text(comment.get("negative_text"), limit=900),
            full_text=_clean_text(comment.get("full_text"), limit=900),
        )
    return {
        "comment_id": _clean_text(comment.get("comment_id"), limit=80),
        "platform": _clean_text(comment.get("platform"), limit=80),
        "date": _clean_text(comment.get("date"), limit=80),
        "model_name": _clean_text(comment.get("model_name"), limit=160),
        "raw_comment": raw_comment,
    }


def _raw_comment_for_fact(comment: dict[str, Any]) -> dict[str, Any]:
    raw_comment = comment.get("raw_comment")
    if isinstance(raw_comment, dict):
        return raw_comment
    return _standard_raw_comment(
        positive_text=_clean_text(comment.get("positive_text"), limit=900),
        negative_text=_clean_text(comment.get("negative_text"), limit=900),
        full_text=_clean_text(comment.get("full_text"), limit=900),
    )


def _append_unique(values: list[str], value: str, *, limit: int) -> None:
    text = _clean_text(value, limit=limit)
    if text and text not in values:
        values.append(text)


def _local_keywords_for_fact(*, text: str, section_facts: dict[str, str]) -> list[str]:
    keywords: list[str] = []
    for section_key in section_facts:
        label = SECTION_KEY_TO_LABEL.get(section_key)
        if label:
            _append_unique(keywords, label, limit=20)
    for keyword, needles in LOCAL_KEYWORD_RULES:
        if any(needle in text for needle in needles):
            _append_unique(keywords, keyword, limit=20)
    return keywords[:10] or ["整体体验"]


def _analysis_fact_for_llm(comment: dict[str, Any]) -> dict[str, Any]:
    raw_comment = _raw_comment_for_fact(comment)
    positive = _clean_text(raw_comment.get("positive_text") or comment.get("positive_text"), limit=MAX_FACT_SENTIMENT_CHARS)
    negative = _clean_text(raw_comment.get("negative_text") or comment.get("negative_text"), limit=MAX_FACT_SENTIMENT_CHARS)
    full_text = _clean_text(raw_comment.get("full_text") or comment.get("full_text"), limit=220)
    raw_sections = raw_comment.get("sections") if isinstance(raw_comment.get("sections"), dict) else {}
    keyword_source = " ".join(
        [
            positive,
            negative,
            full_text,
            *[_clean_text(value, limit=80) for value in raw_sections.values()],
        ]
    )

    section_facts: dict[str, str] = {}
    if positive:
        section_facts["positive"] = positive
    if negative:
        section_facts["negative"] = negative
    topic_count = 0
    for key, value in raw_sections.items():
        if topic_count >= MAX_FACT_TOPIC_SECTIONS:
            break
        text = _clean_text(value, limit=MAX_FACT_SECTION_CHARS)
        if text:
            section_facts[_clean_text(key, limit=40)] = text
            topic_count += 1

    compact_parts: list[str] = []
    if positive:
        compact_parts.append(f"最满意：{positive}")
    if negative:
        compact_parts.append(f"最不满意：{negative}")
    for key, value in section_facts.items():
        if key in {"positive", "negative"}:
            continue
        label = SECTION_KEY_TO_LABEL.get(key, key)
        compact_parts.append(f"{label}：{value}")
    if not compact_parts and full_text:
        compact_parts.append(f"评价：{full_text}")
    elif full_text and full_text not in "；".join(compact_parts):
        compact_parts.append(f"补充：{full_text[:80]}")

    compact_text = _clean_text("；".join(compact_parts), limit=MAX_FACT_TEXT_CHARS)
    polarity = "mixed" if positive and negative else "positive" if positive else "negative" if negative else "neutral"
    return {
        "comment_id": _clean_text(comment.get("comment_id"), limit=80),
        "platform": _clean_text(comment.get("platform"), limit=80),
        "date": _clean_text(comment.get("date"), limit=80),
        "model_name": _clean_text(comment.get("model_name"), limit=160),
        "polarity": polarity,
        "local_keywords": _local_keywords_for_fact(text=keyword_source, section_facts=section_facts)[:5],
        "compact_text": compact_text,
        "section_facts": section_facts,
    }


def _build_analysis_facts(comments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [_analysis_fact_for_llm(comment) for comment in comments]


def _build_batch_prompt(*, model_name: str, batch_index: int, total_batches: int, comments: list[dict[str, Any]]) -> str:
    normalized_comments = [_comment_for_llm(comment) for comment in comments]
    return (
        "你是汽车口碑分析Agent。只根据输入的脱敏原评论JSON分析，不引入外部资料。\n"
        "只返回严格JSON，不要Markdown，不要解释。JSON 字符串内部不要使用英文双引号，引用短语请用中文引号「」。\n"
        "JSON schema: {\"themes\":[{\"direction\":\"positive|negative\",\"term\":\"主题词\",\"count\":数字,"
        "\"summary\":\"摘要\",\"evidence_ids\":[\"comment_id\"]}],\"suggestions\":[{\"direction\":\"方向\",\"text\":\"建议\"}],"
        "\"platform_notes\":[{\"platform\":\"平台\",\"summary\":\"差异\"}],\"boss_brief\":[\"一句话\"]}。\n"
        "证据只允许填写输入中的 comment_id，不要在 JSON 中复制原评论原句。\n"
        f"车型：{model_name}；批次：{batch_index}/{total_batches}。\n"
        "评论白名单字段如下。raw_comment 是规则拆分后的标准原评论JSON，sections 只来自明确分项标题：\n"
        + json.dumps(normalized_comments, ensure_ascii=False, default=_json_default)
    )


def _build_fact_batch_prompt(*, model_name: str, batch_index: int, total_batches: int, facts: list[dict[str, Any]]) -> str:
    return (
        "你是汽车口碑分析Agent。只根据输入的脱敏评论facts分析，不引入外部资料。\n"
        "只返回严格JSON object，不要Markdown，不要解释。JSON 字符串内部不要使用英文双引号，引用短语请用中文引号「」。\n"
        "JSON schema: {\"batch\":\"1/3\",\"themes\":[{\"direction\":\"positive|negative\",\"term\":\"主题词\","
        "\"count\":数字,\"summary\":\"80字内摘要\",\"keywords\":[\"词\"],\"evidence_ids\":[\"comment_id\"],"
        "\"platform_counts\":{\"汽车之家\":0,\"懂车帝\":0}}],\"platform_notes\":[{\"platform\":\"平台\",\"summary\":\"差异\"}]}。\n"
        "每个 theme 的 evidence_ids 只能使用输入中的 comment_id；不要复制原评论长文本；建议和老板口径留给最终汇总。\n"
        f"车型：{model_name}；批次：{batch_index}/{total_batches}。\n"
        "评论facts如下：\n"
        + json.dumps(facts, ensure_ascii=False, default=_json_default)
    )


def _batch_items_by_prompt_budget(
    values: list[dict[str, Any]],
    *,
    target_bytes: int,
    prompt_builder: Any,
) -> list[list[dict[str, Any]]]:
    budget = max(target_bytes, 1_000)
    batches: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    for item in values:
        candidate = [*current, item]
        prompt_bytes = len(prompt_builder(len(batches) + 1, 999, candidate).encode("utf-8"))
        if current and prompt_bytes > budget:
            batches.append(current)
            current = [item]
        else:
            current = candidate
    if current:
        batches.append(current)

    while True:
        total = max(len(batches), 1)
        split_index = next(
            (
                index
                for index, batch in enumerate(batches)
                if len(batch) > 1 and len(prompt_builder(index + 1, total, batch).encode("utf-8")) > budget
            ),
            None,
        )
        if split_index is None:
            return batches
        batch = batches.pop(split_index)
        midpoint = max(1, len(batch) // 2)
        batches.insert(split_index, batch[midpoint:])
        batches.insert(split_index, batch[:midpoint])


def _local_batch_payload(*, batch_index: int, total_batches: int, comments: list[dict[str, str]], reason: str) -> dict[str, Any]:
    evidence_ids = [
        _clean_text(comment.get("comment_id"), limit=80)
        for comment in comments
        if _clean_text(comment.get("comment_id"), limit=80)
    ]
    positive_snippets: list[str] = []
    negative_snippets: list[str] = []
    platform_counter: Counter[str] = Counter()
    for comment in comments:
        platform = _clean_text(comment.get("platform"), limit=80)
        if platform:
            platform_counter[platform] += 1
        positive = _clean_text(comment.get("positive_text"), limit=180)
        negative = _clean_text(comment.get("negative_text"), limit=180)
        full_text = _clean_text(comment.get("full_text"), limit=220)
        if positive:
            positive_snippets.append(positive)
        elif full_text:
            positive_snippets.append(full_text)
        if negative:
            negative_snippets.append(negative)

    fallback_evidence = evidence_ids[:8] or [f"local_batch_{batch_index:03d}"]
    positive_summary = "；".join(positive_snippets[:4]) or "该批 Hermes 输出失败，使用脱敏评论做本地批次兜底。"
    negative_summary = "；".join(negative_snippets[:4]) or "该批未稳定识别独立槽点，建议结合最终汇总继续判断。"
    platform_summary = "；".join(f"{platform} {count} 条" for platform, count in platform_counter.items())
    return {
        "batch": f"{batch_index}/{total_batches}",
        "themes": [
            {
                "direction": "positive",
                "term": "本地批次兜底",
                "count": max(len(positive_snippets), 1),
                "summary": positive_summary,
                "evidence_ids": fallback_evidence,
            },
            {
                "direction": "negative",
                "term": "本地槽点",
                "count": max(len(negative_snippets), 1),
                "summary": negative_summary,
                "evidence_ids": fallback_evidence,
            },
        ],
        "suggestions": [
            {
                "direction": "稳定性",
                "text": f"批次 {batch_index}/{total_batches} 的 Hermes 输出未能解析，已用本地规则兜底；建议优先复核该批证据 id。",
            }
        ],
        "platform_notes": [
            {
                "platform": "本地批次兜底",
                "summary": platform_summary or f"批次 {batch_index}/{total_batches} 共 {len(comments)} 条脱敏评论。",
            }
        ],
        "boss_brief": [f"批次 {batch_index}/{total_batches} 使用本地批次兜底，原因：{_clean_text(reason, limit=120)}"],
    }


def _count_value(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _compact_evidence_ids(value: Any, *, limit: int = 1) -> list[str]:
    ids: list[str] = []
    for item in _as_list(value):
        text = _clean_text(item, limit=80)
        if text:
            ids.append(text)
        if len(ids) >= limit:
            break
    return ids


def _compact_themes(value: Any) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []
    for direction in ("positive", "negative"):
        themes = [
            item
            for item in _as_list(value)
            if isinstance(item, dict) and _clean_text(item.get("direction"), limit=40) == direction
        ]
        themes.sort(key=lambda item: _count_value(item.get("count")), reverse=True)
        for item in themes[:MAX_AGGREGATE_THEMES_PER_DIRECTION]:
            selected.append(
                {
                    "direction": direction,
                    "term": _clean_text(item.get("term"), limit=50),
                    "count": max(_count_value(item.get("count")), 1),
                    "summary": _clean_text(item.get("summary") or item.get("description"), limit=80),
                    "evidence_ids": _compact_evidence_ids(item.get("evidence_ids")),
                }
            )
    return [item for item in selected if item["term"]]


def _compact_batch_payloads(batch_payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    compacted: list[dict[str, Any]] = []
    for index, payload in enumerate(batch_payloads, start=1):
        if not isinstance(payload, dict):
            continue
        suggestions = []
        for item in _as_list(payload.get("suggestions"))[:MAX_AGGREGATE_SUGGESTIONS]:
            if not isinstance(item, dict):
                continue
            text = _clean_text(item.get("text"), limit=80)
            if text:
                suggestions.append({"direction": _clean_text(item.get("direction"), limit=50), "text": text})

        platform_notes = []
        for item in _as_list(payload.get("platform_notes"))[:MAX_AGGREGATE_PLATFORM_NOTES]:
            if not isinstance(item, dict):
                continue
            summary = _clean_text(item.get("summary"), limit=80)
            if summary:
                platform_notes.append({"platform": _clean_text(item.get("platform"), limit=50), "summary": summary})

        compacted.append(
            {
                "batch": _clean_text(payload.get("batch"), limit=40) or str(index),
                "themes": _compact_themes(payload.get("themes")),
                "suggestions": suggestions,
                "platform_notes": platform_notes,
                "boss_brief": [
                    _clean_text(item, limit=80)
                    for item in _as_list(payload.get("boss_brief"))[:1]
                    if _clean_text(item, limit=80)
                ],
            }
        )
    return compacted


def _json_size(value: Any) -> int:
    return len(json.dumps(value, ensure_ascii=False, default=_json_default).encode("utf-8"))


def _fit_aggregate_payloads_to_command_limit(compacted_payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if _json_size(compacted_payloads) <= MAX_AGGREGATE_PAYLOAD_BYTES:
        return compacted_payloads

    reduced: list[dict[str, Any]] = []
    for payload in compacted_payloads:
        themes = []
        for theme in _as_list(payload.get("themes"))[:4]:
            if not isinstance(theme, dict):
                continue
            themes.append(
                {
                    "direction": _clean_text(theme.get("direction"), limit=20),
                    "term": _clean_text(theme.get("term"), limit=40),
                    "count": _count_value(theme.get("count")),
                    "summary": _clean_text(theme.get("summary"), limit=45),
                    "evidence_ids": _compact_evidence_ids(theme.get("evidence_ids"), limit=1),
                }
            )
        reduced.append({"batch": _clean_text(payload.get("batch"), limit=20), "themes": themes})
    if _json_size(reduced) <= MAX_AGGREGATE_PAYLOAD_BYTES:
        return reduced

    fitted: list[dict[str, Any]] = []
    for payload in reduced:
        candidate = [*fitted, payload]
        if _json_size(candidate) > MAX_AGGREGATE_PAYLOAD_BYTES:
            break
        fitted = candidate
    return fitted or reduced[:1]


def _build_aggregate_prompt(*, model_name: str, comments: list[dict[str, str]], batch_payloads: list[dict[str, Any]]) -> str:
    sample_counts = Counter(comment["platform"] for comment in comments)
    compacted_payloads = _fit_aggregate_payloads_to_command_limit(_compact_batch_payloads(batch_payloads))
    return (
        "你是汽车口碑分析Agent。请归并各批分析结果，生成最终结果。只返回严格JSON，不要Markdown。\n"
        "JSON 字符串内部不要使用英文双引号，引用短语请用中文引号「」。\n"
        "JSON schema: {\"headline\":\"一句话结论\",\"executive_summary\":\"摘要\","
        "\"strength_blocks\":[{\"title\":\"核心好评\",\"summary\":\"...\",\"evidence_ids\":[\"...\"]}],"
        "\"weakness_blocks\":[{\"title\":\"核心槽点\",\"summary\":\"...\",\"evidence_ids\":[\"...\"]}],"
        "\"platform_difference_blocks\":[{\"title\":\"方向\",\"summary\":\"...\",\"evidence_ids\":[\"...\"]}],"
        "\"action_blocks\":[{\"title\":\"建议方向\",\"summary\":\"...\",\"evidence_ids\":[\"...\"]}],"
        "\"boss_brief\":[\"...\"],\"keyword_rankings\":{\"positive\":[{\"term\":\"...\",\"count\":1}],"
        "\"negative\":[{\"term\":\"...\",\"count\":1}]},\"qa_chunks\":[{\"chunk_id\":\"...\",\"source_type\":\"hermes_evidence\","
        "\"text\":\"...\",\"tags\":[\"...\"],\"metadata\":{\"source\":\"hermes\"}}],"
        "\"compare_rows\":[{\"方向\":\"...\",\"汽车之家_优势提及\":\"0\",\"汽车之家_槽点提及\":\"0\",\"懂车帝_优势提及\":\"0\",\"懂车帝_槽点提及\":\"0\"}],"
        "\"opportunity_rows\":[{\"类型\":\"改进\",\"方向\":\"...\",\"建议\":\"...\"}]}。\n"
        f"车型：{model_name}；样本数：汽车之家 {sample_counts.get(PLATFORM_AUTOHOME, 0)} 条，懂车帝 {sample_counts.get(PLATFORM_DCD, 0)} 条。\n"
        "批次结果如下：\n"
        + json.dumps(compacted_payloads, ensure_ascii=False, default=_json_default)
    )


def _call_hermes(
    prompt: str,
    *,
    hermes_command: str,
    env: dict[str, str],
    debug_dir: Path | None = None,
    call_label: str = "hermes",
) -> Any:
    command_path = _hermes_command_path(hermes_command)
    if not command_path:
        raise RuntimeError("hermes command not found")

    env = _write_hermes_config(env)
    provider, model, env = _runtime_provider(env)
    timeout = int(env.get("HERMES_TIMEOUT_SECONDS") or "180")
    json_retries = max(0, int(env.get("HERMES_JSON_RETRIES") or "1"))
    parse_error: Exception | None = None
    current_prompt = prompt
    for attempt in range(json_retries + 1):
        attempt_no = attempt + 1
        try:
            completed = subprocess.run(
                [
                    command_path,
                    "chat",
                    "--provider",
                    provider,
                    "--model",
                    model,
                    "-Q",
                    "-q",
                    current_prompt,
                ],
                capture_output=True,
                text=True,
                check=False,
                timeout=timeout,
                env=env,
            )
        except subprocess.TimeoutExpired as exc:
            stdout = _process_output_text(exc.stdout)
            stderr = _process_output_text(exc.stderr)
            if stdout:
                _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="stdout", content=stdout)
            if stderr:
                _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="stderr", content=stderr)
            _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="timeout", content=f"timed out after {timeout} seconds")
            raise RuntimeError(f"hermes_timeout:{_safe_debug_label(call_label)}:{timeout}s") from exc
        _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="stdout", content=completed.stdout)
        if completed.stderr:
            _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="stderr", content=completed.stderr)
        if completed.returncode != 0:
            raise RuntimeError(completed.stderr.strip() or completed.stdout.strip() or "hermes failed")
        try:
            return _extract_json(completed.stdout)
        except Exception as exc:
            parse_error = exc
            _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="parse_error", content=str(exc))
            if attempt < json_retries:
                current_prompt = _build_json_repair_prompt(raw_response=completed.stdout, parse_error=exc)

    raise ValueError(f"hermes_invalid_json:{parse_error}") from parse_error


def _new_llm_metrics(*, mode: str, batch_model: str, aggregate_model: str, comment_count: int, fact_count: int) -> dict[str, Any]:
    return {
        "mode": mode,
        "models": {"batch": batch_model, "aggregate": aggregate_model},
        "comment_count": comment_count,
        "fact_count": fact_count,
        "batch_count": 0,
        "calls": {"batch": 0, "aggregate": 0},
        "prompt_bytes": {"batch": 0, "aggregate": 0},
        "output_bytes": {"batch": 0, "aggregate": 0},
        "retry_count": {"batch": 0, "aggregate": 0},
        "parse_errors": {"batch": 0, "aggregate": 0},
        "fallbacks": {"batch": [], "aggregate": None},
        "durations_ms": {"batch": 0, "aggregate": 0, "total": 0},
        "wall_durations_ms": {"batch": 0, "aggregate": 0, "total": 0},
        "duration_note": "durations_ms sums per-call durations; wall_durations_ms records elapsed wall time.",
    }


def _record_llm_metrics(
    metrics: dict[str, Any] | None,
    *,
    stage: str,
    prompt: str,
    output: str,
    retries: int,
    parse_errors: int,
    started_at: float,
) -> None:
    if metrics is None:
        return
    duration_ms = int((time.monotonic() - started_at) * 1000)
    with LLM_METRICS_LOCK:
        calls = metrics.setdefault("calls", {})
        prompt_bytes = metrics.setdefault("prompt_bytes", {})
        output_bytes = metrics.setdefault("output_bytes", {})
        retry_count = metrics.setdefault("retry_count", {})
        parse_error_counts = metrics.setdefault("parse_errors", {})
        durations = metrics.setdefault("durations_ms", {})
        calls[stage] = int(calls.get(stage) or 0) + 1
        prompt_bytes[stage] = int(prompt_bytes.get(stage) or 0) + len(prompt.encode("utf-8"))
        output_bytes[stage] = int(output_bytes.get(stage) or 0) + len(output.encode("utf-8"))
        retry_count[stage] = int(retry_count.get(stage) or 0) + retries
        parse_error_counts[stage] = int(parse_error_counts.get(stage) or 0) + parse_errors
        durations[stage] = int(durations.get(stage) or 0) + duration_ms


def _set_llm_wall_duration(metrics: dict[str, Any] | None, stage: str, started_at: float) -> None:
    if metrics is None:
        return
    duration_ms = int((time.monotonic() - started_at) * 1000)
    with LLM_METRICS_LOCK:
        wall_durations = metrics.setdefault("wall_durations_ms", {})
        wall_durations[stage] = duration_ms


def _aggregate_hard_timeout_seconds(env: dict[str, str]) -> int:
    raw = (env.get("HERMES_AGGREGATE_TIMEOUT_SECONDS") or env.get("HERMES_TIMEOUT_SECONDS") or "180").strip()
    try:
        return max(int(raw), 1)
    except (TypeError, ValueError):
        return 180


def _can_use_signal_hard_timeout() -> bool:
    return (
        current_thread() is main_thread()
        and hasattr(signal, "SIGALRM")
        and hasattr(signal, "ITIMER_REAL")
        and hasattr(signal, "setitimer")
    )


def _call_aggregate_llm_json(
    prompt: str,
    *,
    hermes_command: str,
    env: dict[str, str],
    model: str,
    debug_dir: Path | None,
    call_label: str,
    metrics: dict[str, Any] | None,
    metric_stage: str,
) -> Any:
    hard_timeout_seconds = _aggregate_hard_timeout_seconds(env)
    started_at = time.monotonic()

    if _hermes_llm_mode(env) != "api" or not _can_use_signal_hard_timeout():
        try:
            return _call_llm_json(
                prompt,
                hermes_command=hermes_command,
                env=env,
                model=model,
                debug_dir=debug_dir,
                call_label=call_label,
                metrics=metrics,
                metric_stage=metric_stage,
            )
        finally:
            _set_llm_wall_duration(metrics, metric_stage, started_at)

    old_handler = signal.getsignal(signal.SIGALRM)
    old_timer = signal.getitimer(signal.ITIMER_REAL)

    def raise_hard_timeout(_signum: int, _frame: Any) -> None:
        raise _AggregateHardTimeout(f"aggregate_hard_timeout:{_safe_debug_label(call_label)}:{hard_timeout_seconds}s")

    signal.signal(signal.SIGALRM, raise_hard_timeout)
    signal.setitimer(signal.ITIMER_REAL, hard_timeout_seconds)
    try:
        return _call_llm_json(
            prompt,
            hermes_command=hermes_command,
            env=env,
            model=model,
            debug_dir=debug_dir,
            call_label=call_label,
            metrics=metrics,
            metric_stage=metric_stage,
        )
    except _AggregateHardTimeout as exc:
        reason = str(exc) or f"aggregate_hard_timeout:{_safe_debug_label(call_label)}:{hard_timeout_seconds}s"
        _write_hermes_debug(debug_dir, label=call_label, attempt=1, kind="timeout", content=f"hard timed out after {hard_timeout_seconds} seconds")
        _record_llm_metrics(
            metrics,
            stage=metric_stage,
            prompt=prompt,
            output="",
            retries=0,
            parse_errors=0,
            started_at=started_at,
        )
        raise RuntimeError(reason) from exc
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, old_handler)
        if old_timer[0] > 0 or old_timer[1] > 0:
            signal.setitimer(signal.ITIMER_REAL, old_timer[0], old_timer[1])
        _set_llm_wall_duration(metrics, metric_stage, started_at)


def _call_deepseek_json(
    prompt: str,
    *,
    env: dict[str, str],
    model: str,
    debug_dir: Path | None = None,
    call_label: str = "deepseek",
    metrics: dict[str, Any] | None = None,
    metric_stage: str = "batch",
) -> Any:
    api_key = (env.get("LLM_API_KEY") or env.get("DEEPSEEK_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError("deepseek_disabled:missing_llm_api_key")
    base_url = _deepseek_base_url(env)
    timeout = _int_env(env, "HERMES_TIMEOUT_SECONDS", 180)
    json_retries = max(0, _int_env(env, "HERMES_JSON_RETRIES", 1, minimum=0))
    retry_base_seconds = float(env.get("HERMES_RETRY_BASE_SECONDS") or "0.5")
    attempts = json_retries + 1
    parse_error: Exception | None = None
    last_error: Exception | None = None
    output_text = ""
    started_at = time.monotonic()
    parse_errors = 0
    retries_used = 0

    request_body = {
        "model": _normalize_deepseek_model(model),
        "messages": [
            {
                "role": "system",
                "content": "你是汽车口碑分析Agent。必须只输出一个合法 JSON object，不要输出 Markdown 或解释文字。",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.1,
        "response_format": {"type": "json_object"},
    }
    headers = {"Authorization": f"Bearer {api_key}"}

    for attempt in range(attempts):
        attempt_no = attempt + 1
        try:
            response = requests.post(
                f"{base_url}/chat/completions",
                headers=headers,
                json=request_body,
                timeout=timeout,
            )
        except requests.Timeout as exc:
            last_error = exc
            _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="timeout", content=f"timed out after {timeout} seconds")
            if attempt < attempts - 1:
                retries_used += 1
                if retry_base_seconds > 0:
                    time.sleep(retry_base_seconds * (2**attempt))
                continue
            _record_llm_metrics(metrics, stage=metric_stage, prompt=prompt, output=output_text, retries=retries_used, parse_errors=parse_errors, started_at=started_at)
            raise RuntimeError(f"deepseek_timeout:{_safe_debug_label(call_label)}:{timeout}s") from exc
        except requests.RequestException as exc:
            last_error = exc
            _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="request_error", content=str(exc))
            if attempt < attempts - 1:
                retries_used += 1
                if retry_base_seconds > 0:
                    time.sleep(retry_base_seconds * (2**attempt))
                continue
            _record_llm_metrics(metrics, stage=metric_stage, prompt=prompt, output=output_text, retries=retries_used, parse_errors=parse_errors, started_at=started_at)
            raise RuntimeError(f"deepseek_request_error:{_safe_debug_label(call_label)}") from exc

        output_text = response.text or ""
        _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="stdout", content=output_text)
        status_code = int(getattr(response, "status_code", 0) or 0)
        if status_code in DEEPSEEK_RETRYABLE_STATUS_CODES and attempt < attempts - 1:
            retries_used += 1
            if retry_base_seconds > 0:
                time.sleep(retry_base_seconds * (2**attempt))
            continue
        if status_code >= 400:
            _record_llm_metrics(metrics, stage=metric_stage, prompt=prompt, output=output_text, retries=retries_used, parse_errors=parse_errors, started_at=started_at)
            raise RuntimeError(f"deepseek_http:{status_code}:{_safe_debug_label(call_label)}")

        try:
            response_payload = response.json()
            content = response_payload["choices"][0]["message"]["content"]
            if not _clean_text(content, limit=20):
                raise ValueError("empty response")
            payload = _extract_json(str(content))
            _record_llm_metrics(metrics, stage=metric_stage, prompt=prompt, output=str(content), retries=retries_used, parse_errors=parse_errors, started_at=started_at)
            return payload
        except Exception as exc:
            parse_error = exc
            parse_errors += 1
            _write_hermes_debug(debug_dir, label=call_label, attempt=attempt_no, kind="parse_error", content=str(exc))
            if attempt < attempts - 1:
                retries_used += 1
                if retry_base_seconds > 0:
                    time.sleep(retry_base_seconds * (2**attempt))
                continue

    _record_llm_metrics(metrics, stage=metric_stage, prompt=prompt, output=output_text, retries=retries_used, parse_errors=parse_errors, started_at=started_at)
    if parse_error is not None:
        raise ValueError(f"deepseek_invalid_json:{parse_error}") from parse_error
    raise RuntimeError(f"deepseek_failed:{last_error}") from last_error


def _call_llm_json(
    prompt: str,
    *,
    hermes_command: str,
    env: dict[str, str],
    model: str,
    debug_dir: Path | None,
    call_label: str,
    metrics: dict[str, Any] | None,
    metric_stage: str,
) -> Any:
    if _hermes_llm_mode(env) == "api":
        return _call_deepseek_json(
            prompt,
            env=env,
            model=model,
            debug_dir=debug_dir,
            call_label=call_label,
            metrics=metrics,
            metric_stage=metric_stage,
        )

    cli_env = dict(env)
    cli_env["LLM_MODEL_REPORT"] = model
    started_at = time.monotonic()
    try:
        payload = _call_hermes(prompt, hermes_command=hermes_command, env=cli_env, debug_dir=debug_dir, call_label=call_label)
        _record_llm_metrics(metrics, stage=metric_stage, prompt=prompt, output=json.dumps(payload, ensure_ascii=False, default=_json_default), retries=0, parse_errors=0, started_at=started_at)
        return payload
    except Exception:
        _record_llm_metrics(metrics, stage=metric_stage, prompt=prompt, output="", retries=0, parse_errors=1, started_at=started_at)
        raise


def _as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _block(title: str, summary: str, evidence_id: str) -> dict[str, Any]:
    return {"title": title, "summary": summary, "evidence_ids": [evidence_id]}


def _rankings_from_themes(batch_payloads: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    counters: dict[str, Counter[str]] = {"positive": Counter(), "negative": Counter()}
    for payload in batch_payloads:
        for theme in _as_list(payload.get("themes")):
            if not isinstance(theme, dict):
                continue
            direction = theme.get("direction")
            if direction not in counters:
                continue
            term = _clean_text(theme.get("term"), limit=80)
            if not term:
                continue
            count = int(theme.get("count") or 1)
            counters[direction][term] += max(count, 1)
    return {
        direction: [{"term": term, "count": count} for term, count in counter.most_common(10)]
        for direction, counter in counters.items()
    }


def _normalize_rankings(payload: dict[str, Any], batch_payloads: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    rankings = payload.get("keyword_rankings") if isinstance(payload.get("keyword_rankings"), dict) else {}
    derived = _rankings_from_themes(batch_payloads)
    normalized: dict[str, list[dict[str, Any]]] = {}
    for direction in ("positive", "negative"):
        values = _as_list(rankings.get(direction))
        items: list[dict[str, Any]] = []
        for item in values:
            if not isinstance(item, dict):
                continue
            term = _clean_text(item.get("term"), limit=80)
            if not term:
                continue
            items.append({"term": term, "count": int(item.get("count") or item.get("weight") or 1)})
        normalized[direction] = items[:10] or derived[direction]
    return normalized


def _normalize_blocks(value: Any, *, fallback_title: str, fallback_summary: str, evidence_id: str) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    for item in _as_list(value):
        if not isinstance(item, dict):
            continue
        title = _clean_generated_text(item.get("title"), limit=80) or fallback_title
        summary = _clean_generated_text(item.get("summary"), limit=500)
        if not summary:
            continue
        evidence_ids = [str(eid) for eid in _as_list(item.get("evidence_ids")) if str(eid).strip()]
        blocks.append({"title": title, "summary": summary, "evidence_ids": evidence_ids or [evidence_id]})
    return blocks or [_block(fallback_title, fallback_summary, evidence_id)]


def _summary_from_rankings(items: list[dict[str, Any]]) -> str:
    terms = [str(item.get("term")) for item in items[:5] if item.get("term")]
    return "、".join(terms) if terms else "样本不足，暂未形成稳定主题"


def _evidence_ids_from_themes(batch_payloads: list[dict[str, Any]], direction: str, *, limit: int = 8) -> list[str]:
    evidence_ids: list[str] = []
    for payload in batch_payloads:
        for theme in _as_list(payload.get("themes")):
            if not isinstance(theme, dict) or theme.get("direction") != direction:
                continue
            for evidence_id in _compact_evidence_ids(theme.get("evidence_ids"), limit=3):
                if evidence_id not in evidence_ids:
                    evidence_ids.append(evidence_id)
                if len(evidence_ids) >= limit:
                    return evidence_ids
    return evidence_ids


def _local_aggregate_payload(
    *,
    model_name: str,
    comments: list[dict[str, str]],
    batch_payloads: list[dict[str, Any]],
) -> dict[str, Any]:
    rankings = _rankings_from_themes(batch_payloads)
    positive_summary = _summary_from_rankings(rankings["positive"])
    negative_summary = _summary_from_rankings(rankings["negative"])
    headline = f"{model_name} 口碑核心好评集中在{positive_summary}，主要槽点集中在{negative_summary}。"

    suggestions: list[dict[str, str]] = []
    platform_notes: list[dict[str, str]] = []
    boss_brief: list[str] = []
    for payload in batch_payloads:
        for item in _as_list(payload.get("suggestions")):
            if isinstance(item, dict):
                text = _clean_generated_text(item.get("text"), limit=300)
                if text:
                    suggestions.append({"direction": _clean_generated_text(item.get("direction"), limit=80), "text": text})
        for item in _as_list(payload.get("platform_notes")):
            if isinstance(item, dict):
                summary = _clean_generated_text(item.get("summary"), limit=320)
                if summary:
                    platform_notes.append({"platform": _clean_generated_text(item.get("platform"), limit=80), "summary": summary})
        for line in _as_list(payload.get("boss_brief")):
            text = _clean_generated_text(line, limit=240)
            if text:
                boss_brief.append(text)

    action_blocks = [
        {
            "title": suggestion.get("direction") or "产品建议",
            "summary": suggestion["text"],
            "evidence_ids": [f"hermes.suggestion.{index}"],
        }
        for index, suggestion in enumerate(suggestions[:5], start=1)
    ]
    platform_difference_blocks = [
        {
            "title": note.get("platform") or "平台差异",
            "summary": note["summary"],
            "evidence_ids": [f"hermes.platform.{index}"],
        }
        for index, note in enumerate(platform_notes[:4], start=1)
    ]
    opportunity_rows = [
        {
            "类型": "改进",
            "方向": suggestion.get("direction") or "产品体验",
            "建议": suggestion["text"],
        }
        for suggestion in suggestions[:8]
    ]
    compare_rows = [
        {
            "方向": item["term"],
            "汽车之家_优势提及": str(item["count"]),
            "汽车之家_槽点提及": "0",
            "懂车帝_优势提及": "0",
            "懂车帝_槽点提及": "0",
        }
        for item in rankings["positive"][:5]
    ]
    compare_rows.extend(
        {
            "方向": item["term"],
            "汽车之家_优势提及": "0",
            "汽车之家_槽点提及": str(item["count"]),
            "懂车帝_优势提及": "0",
            "懂车帝_槽点提及": "0",
        }
        for item in rankings["negative"][:5]
    )

    return {
        "headline": headline,
        "executive_summary": f"基于 {len(comments)} 条双平台脱敏原评论的 Hermes 批次分析归并。{headline}",
        "strength_blocks": [
            {
                "title": "核心好评",
                "summary": positive_summary,
                "evidence_ids": _evidence_ids_from_themes(batch_payloads, "positive") or ["hermes.strengths"],
            }
        ],
        "weakness_blocks": [
            {
                "title": "核心槽点",
                "summary": negative_summary,
                "evidence_ids": _evidence_ids_from_themes(batch_payloads, "negative") or ["hermes.weaknesses"],
            }
        ],
        "platform_difference_blocks": platform_difference_blocks,
        "action_blocks": action_blocks,
        "boss_brief": boss_brief[:3],
        "keyword_rankings": rankings,
        "qa_chunks": [],
        "compare_rows": compare_rows,
        "opportunity_rows": opportunity_rows,
    }


def _comment_evidence_chunks(comments: list[dict[str, str]], model_name: str) -> list[dict[str, Any]]:
    chunks: list[dict[str, Any]] = []
    for index, comment in enumerate(comments, start=1):
        comment_id = comment.get("comment_id") or f"comment_{index:04d}"
        parts = [
            f"平台：{comment.get('platform', '')}",
            f"日期：{comment.get('date', '')}",
            f"车型：{comment.get('model_name') or model_name}",
        ]
        if comment.get("positive_text"):
            parts.append(f"最满意：{comment['positive_text']}")
        if comment.get("negative_text"):
            parts.append(f"最不满意：{comment['negative_text']}")
        if comment.get("full_text"):
            parts.append(f"评价全文：{comment['full_text']}")
        text = _clean_generated_text("；".join(part for part in parts if part), limit=1600)
        if not text:
            continue
        chunks.append(
            {
                "chunk_id": comment_id,
                "source_type": "comment_evidence",
                "text": text,
                "tags": [comment.get("platform", ""), model_name, "原评论证据"],
                "metadata": {
                    "source": "hermes_whitelisted_comment",
                    "platform": comment.get("platform", ""),
                    "date": comment.get("date", ""),
                    "model_name": comment.get("model_name") or model_name,
                },
            }
        )
    return chunks


def _normalize_qa_chunks(
    payload: dict[str, Any],
    report: dict[str, Any],
    model_name: str,
    comments: list[dict[str, str]] | None = None,
) -> list[dict[str, Any]]:
    chunks: list[dict[str, Any]] = []
    for index, item in enumerate(_as_list(payload.get("qa_chunks")), start=1):
        if not isinstance(item, dict):
            continue
        text = _clean_generated_text(item.get("text"), limit=1000)
        if not text:
            continue
        chunks.append(
            {
                "chunk_id": _clean_generated_text(item.get("chunk_id"), limit=120) or f"hermes_{index}",
                "source_type": _clean_generated_text(item.get("source_type"), limit=80) or "hermes_evidence",
                "text": text,
                "tags": [_clean_generated_text(tag, limit=120) for tag in _as_list(item.get("tags")) if str(tag).strip()],
                "metadata": _replace_legacy_labels(item.get("metadata")) if isinstance(item.get("metadata"), dict) else {"source": "hermes"},
            }
        )
    evidence_chunks = _comment_evidence_chunks(comments or [], model_name)
    if chunks:
        return chunks + evidence_chunks

    chunk_id = 1
    for source_type, blocks in (
        ("strength", report.get("strength_blocks") or []),
        ("weakness", report.get("weakness_blocks") or []),
        ("action", report.get("action_blocks") or []),
        ("comparison", report.get("platform_difference_blocks") or []),
    ):
        for block in blocks:
            text = f"{block.get('title', '')}：{block.get('summary', '')}"
            if _clean_text(text):
                chunks.append(
                    {
                        "chunk_id": f"hermes_{chunk_id}",
                        "source_type": "hermes_evidence",
                        "text": text,
                        "tags": [source_type, model_name],
                        "metadata": {"source": "hermes"},
                    }
                )
                chunk_id += 1
    return chunks + evidence_chunks


def _normalize_hermes_payload(
    payload: dict[str, Any],
    *,
    model_name: str,
    comments: list[dict[str, str]],
    batch_payloads: list[dict[str, Any]],
) -> NormalizedHermesResult:
    rankings = _normalize_rankings(payload, batch_payloads)
    positive_summary = _summary_from_rankings(rankings["positive"])
    negative_summary = _summary_from_rankings(rankings["negative"])

    headline = _clean_generated_text(payload.get("headline"), limit=200) or f"{model_name} 口碑核心好评集中在{positive_summary}，主要槽点集中在{negative_summary}。"
    report = {
        "headline": headline,
        "executive_summary": _clean_generated_text(payload.get("executive_summary"), limit=800) or headline,
        "strength_blocks": _normalize_blocks(
            payload.get("strength_blocks"),
            fallback_title="核心好评",
            fallback_summary=positive_summary,
            evidence_id="hermes.strengths",
        ),
        "weakness_blocks": _normalize_blocks(
            payload.get("weakness_blocks"),
            fallback_title="核心槽点",
            fallback_summary=negative_summary,
            evidence_id="hermes.weaknesses",
        ),
        "platform_difference_blocks": _normalize_blocks(
            payload.get("platform_difference_blocks"),
            fallback_title="平台差异",
            fallback_summary="两平台反馈方向已在关键词和原句证据中归并。",
            evidence_id="hermes.platform",
        ),
        "action_blocks": _normalize_blocks(
            payload.get("action_blocks"),
            fallback_title="产品建议",
            fallback_summary="优先针对高频不满意主题制定产品和传播动作。",
            evidence_id="hermes.actions",
        ),
        "boss_brief": [
            _clean_generated_text(line, limit=200)
            for line in _as_list(payload.get("boss_brief"))
            if _clean_generated_text(line, limit=200)
        ][:3],
    }
    if len(report["boss_brief"]) < 3:
        report["boss_brief"] = [
            report["headline"],
            f"最满意TOP集中在：{positive_summary}。",
            f"最不满意TOP集中在：{negative_summary}。",
        ][:3]

    compare_rows = [_replace_legacy_labels(row) for row in _as_list(payload.get("compare_rows")) if isinstance(row, dict)]
    opportunity_rows = [_replace_legacy_labels(row) for row in _as_list(payload.get("opportunity_rows")) if isinstance(row, dict)]
    one_pager_lines = [
        "双平台口碑一页纸总结",
        report["headline"],
        f"最满意TOP5：{positive_summary}",
        f"最不满意TOP5：{negative_summary}",
        *report["boss_brief"],
    ]
    return NormalizedHermesResult(
        report=report,
        keyword_rankings=rankings,
        qa_chunks=_normalize_qa_chunks(payload, report, model_name, comments),
        one_pager_lines=one_pager_lines,
        compare_rows=compare_rows,
        opportunity_rows=opportunity_rows,
    )


def _write_summary_workbook(
    path: Path,
    *,
    model_name: str,
    comments: list[dict[str, str]],
    result: NormalizedHermesResult,
) -> None:
    counts = Counter(comment["platform"] for comment in comments)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "总览摘要"
    overview.append(["模块", "内容"])
    overview.append(["项目", model_name])
    overview.append(["平台样本", f"汽车之家 {counts.get(PLATFORM_AUTOHOME, 0)} 条；懂车帝 {counts.get(PLATFORM_DCD, 0)} 条"])
    overview.append(["综合一句话", result.report["headline"]])

    compare = workbook.create_sheet("跨平台对比")
    compare.append(["方向", "汽车之家_优势提及", "汽车之家_槽点提及", "懂车帝_优势提及", "懂车帝_槽点提及"])
    for row in result.compare_rows[:8]:
        compare.append(
            [
                _clean_text(row.get("方向"), limit=80) or _clean_text(row.get("title"), limit=80) or "平台差异",
                _clean_text(row.get("汽车之家_优势提及")) or "0",
                _clean_text(row.get("汽车之家_槽点提及")) or "0",
                _clean_text(row.get("懂车帝_优势提及")) or "0",
                _clean_text(row.get("懂车帝_槽点提及")) or "0",
            ]
        )
    if compare.max_row == 1:
        compare.append(["整体", "0", "0", "0", "0"])

    business = workbook.create_sheet("综合业务摘要")
    business.append(["模块", "内容"])
    business.append(["核心好评", result.report["strength_blocks"][0]["summary"]])
    business.append(["核心槽点", result.report["weakness_blocks"][0]["summary"]])
    business.append(["产品建议", result.report["action_blocks"][0]["summary"]])
    business.append(["适合人群", "关注高频好评主题且能接受当前槽点的潜在用户"])

    opportunity = workbook.create_sheet("产品机会点")
    opportunity.append(["类型", "方向", "建议"])
    for row in result.opportunity_rows[:8]:
        opportunity.append(
            [
                _clean_text(row.get("类型"), limit=80) or "改进",
                _clean_text(row.get("方向"), limit=80) or _clean_text(row.get("title"), limit=80) or "产品体验",
                _clean_text(row.get("建议"), limit=500) or _clean_text(row.get("summary"), limit=500),
            ]
        )
    if opportunity.max_row == 1:
        for block in result.report["action_blocks"][:3]:
            opportunity.append(["改进", block["title"], block["summary"]])

    one_pager = workbook.create_sheet("一页纸总结")
    for line in result.one_pager_lines:
        one_pager.append([line])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_terms_workbook(path: Path, rankings: dict[str, list[dict[str, Any]]]) -> None:
    workbook = Workbook()
    positive = workbook.active
    positive.title = "positive_terms"
    positive.append(["term", "weight"])
    for item in rankings.get("positive", []):
        positive.append([item.get("term"), item.get("count", 1)])

    negative = workbook.create_sheet("negative_terms")
    negative.append(["term", "weight"])
    for item in rankings.get("negative", []):
        negative.append([item.get("term"), item.get("count", 1)])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _font_path(font_path: str | None) -> str | None:
    if not font_path:
        return None
    path = Path(font_path)
    return str(path) if path.exists() else None


def _fallback_wordcloud_png(path: Path, title: str, terms: list[dict[str, Any]], font_path: str | None) -> None:
    from PIL import Image, ImageDraw, ImageFont

    path.parent.mkdir(parents=True, exist_ok=True)
    image = Image.new("RGB", (900, 500), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    cjk_font = _font_path(font_path)
    if cjk_font:
        try:
            font = ImageFont.truetype(cjk_font, 28)
        except OSError:
            font = ImageFont.load_default()
    draw.text((40, 40), title, fill=(30, 30, 30), font=font)
    text = "  ".join(str(item.get("term", "")) for item in terms[:20] if item.get("term"))
    draw.text((40, 110), text or "暂无高频词", fill=(70, 70, 70), font=font)
    image.save(path)


def _write_wordclouds(output_dir: Path, *, model_name: str, rankings: dict[str, list[dict[str, Any]]], font_path: str | None) -> list[str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    image_paths = [
        output_dir / f"{model_name}_优点词云.png",
        output_dir / f"{model_name}_槽点词云.png",
    ]
    try:
        from wordcloud import WordCloud

        for direction, image_path in (("positive", image_paths[0]), ("negative", image_paths[1])):
            frequencies = {str(item["term"]): int(item.get("count") or 1) for item in rankings.get(direction, []) if item.get("term")}
            if not frequencies:
                _fallback_wordcloud_png(image_path, image_path.stem, rankings.get(direction, []), font_path)
                continue
            cloud = WordCloud(
                width=900,
                height=500,
                background_color="white",
                font_path=_font_path(font_path),
                collocations=False,
            )
            cloud.generate_from_frequencies(frequencies)
            cloud.to_file(str(image_path))
    except Exception:
        _fallback_wordcloud_png(image_paths[0], f"{model_name} 优点词云", rankings.get("positive", []), font_path)
        _fallback_wordcloud_png(image_paths[1], f"{model_name} 槽点词云", rankings.get("negative", []), font_path)
    return [str(path) for path in image_paths if path.exists()]


def _write_report_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_replace_legacy_labels(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _write_qa_chunks(path: Path, chunks: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(chunks, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_normalized_comments_jsonl(path: Path, comments: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        json.dumps(_comment_for_llm(comment), ensure_ascii=False, default=_json_default)
        for comment in comments
    ]
    path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def _write_analysis_facts_jsonl(path: Path, facts: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [json.dumps(fact, ensure_ascii=False, default=_json_default) for fact in facts]
    path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def _write_llm_metrics(path: Path, metrics: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(metrics, ensure_ascii=False, indent=2, default=_json_default), encoding="utf-8")


def _write_validation(path: Path, *, source: str, degraded: bool) -> None:
    path.write_text(
        json.dumps({"ok": True, "source": source, "degraded": degraded}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _run_checked(command: list[str], *, cwd: Path | None = None) -> subprocess.CompletedProcess[str]:
    completed = subprocess.run(command, cwd=cwd, capture_output=True, text=True, check=False)
    if completed.returncode != 0:
        raise RuntimeError(completed.stderr.strip() or completed.stdout.strip() or f"{command[0]} failed")
    return completed


def _read_table(workbook, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in workbook.sheetnames:
        return []
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    header = [_clean_text(cell) for cell in rows[0]]
    values: list[dict[str, str]] = []
    for row in rows[1:]:
        item = {header[index]: _clean_text(row[index]) for index in range(min(len(header), len(row))) if header[index]}
        if any(item.values()):
            values.append(item)
    return values


def _read_one_pager(workbook) -> list[str]:
    if "一页纸总结" not in workbook.sheetnames:
        return []
    lines: list[str] = []
    for row in workbook["一页纸总结"].iter_rows(values_only=True):
        line = " ".join(_clean_text(value) for value in row if _clean_text(value))
        if line:
            lines.append(line)
    return lines


def _report_from_summary(summary_path: Path, *, model_name: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    workbook = load_workbook(summary_path, data_only=True)
    try:
        overview = {row.get("模块", ""): row.get("内容", "") for row in _read_table(workbook, "总览摘要")}
        business = {row.get("模块", ""): row.get("内容", "") for row in _read_table(workbook, "综合业务摘要")}
        opportunities = _read_table(workbook, "产品机会点")
        one_pager = _read_one_pager(workbook)
    finally:
        workbook.close()

    report = {
        "headline": overview.get("综合一句话") or (one_pager[1] if len(one_pager) > 1 else f"{model_name} 口碑摘要"),
        "executive_summary": one_pager[1] if len(one_pager) > 1 else overview.get("综合一句话", ""),
        "strength_blocks": [_block("核心好评", business.get("核心好评") or business.get("核心卖点", ""), "business.core_strengths")],
        "weakness_blocks": [_block("核心槽点", business.get("核心槽点", ""), "business.core_weaknesses")],
        "platform_difference_blocks": [],
        "action_blocks": [
            _block(row.get("方向") or row.get("类型") or "产品建议", row.get("建议", ""), f"opportunity.{index}")
            for index, row in enumerate(opportunities[:3], start=1)
            if row.get("建议")
        ],
        "boss_brief": [line for line in one_pager[1:4] if line][:3],
    }
    chunks = _normalize_qa_chunks({}, report, model_name)
    return report, chunks


def _run_rule_fallback(
    *,
    autohome_input: Path,
    dcd_input: Path | None,
    summary_output: Path,
    terms_output: Path,
    wordcloud_output_dir: Path,
    final_report_output: Path,
    qa_chunks_output: Path,
    model_name: str,
    progress_file: Path,
    summary_script: Path,
    wordcloud_script: Path,
    single_platform: bool,
    font_path: str | None,
) -> dict[str, Any]:
    progress_file.parent.mkdir(parents=True, exist_ok=True)
    summary_command = [
        sys.executable,
        str(summary_script),
        "--output",
        str(summary_output),
        "--model-name",
        model_name,
        "--progress-file",
        str(progress_file),
    ]
    if single_platform or not dcd_input or not dcd_input.exists():
        single_input = autohome_input if autohome_input.exists() else dcd_input
        if single_input is None:
            single_input = autohome_input
        summary_command.extend(["--input", str(single_input)])
    else:
        summary_command.extend(["--autohome-input", str(autohome_input), "--dcd-input", str(dcd_input)])
    _run_checked(summary_command, cwd=summary_script.parent)

    wordcloud_command = [
        sys.executable,
        str(wordcloud_script),
        "--input",
        str(summary_output),
        "--output-dir",
        str(wordcloud_output_dir),
        "--model-name",
        model_name,
        "--json",
    ]
    if font_path:
        wordcloud_command.extend(["--font-path", font_path])
    completed = _run_checked(wordcloud_command, cwd=wordcloud_script.parent)
    try:
        payload = _extract_json(completed.stdout)
    except Exception:
        payload = {}
    produced_terms = Path(payload.get("excel_path", "")) if isinstance(payload, dict) and payload.get("excel_path") else terms_output
    if produced_terms.exists() and produced_terms.resolve() != terms_output.resolve():
        terms_output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(produced_terms, terms_output)

    report, chunks = _report_from_summary(summary_output, model_name=model_name)
    _write_report_json(final_report_output, report)
    _write_qa_chunks(qa_chunks_output, chunks)
    _write_validation(summary_output.with_suffix(".validation.json"), source="rule-fallback", degraded=True)
    return {
        "summary_path": str(summary_output),
        "terms_path": str(terms_output),
        "final_report_path": str(final_report_output),
        "qa_chunks_path": str(qa_chunks_output),
        "image_paths": payload.get("image_paths", []) if isinstance(payload, dict) else [],
    }


def _source_name(env: dict[str, str], *, batch_fallbacks: list[dict[str, Any]], aggregate_local: bool = False) -> str:
    prefix = "hermes-deepseek-api" if _hermes_llm_mode(env) == "api" else "hermes"
    if aggregate_local:
        return f"{prefix}-local-aggregate"
    if batch_fallbacks:
        return f"{prefix}-partial-local-batch"
    return prefix


def _comments_by_id(comments: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {_clean_text(comment.get("comment_id"), limit=80): comment for comment in comments if _clean_text(comment.get("comment_id"), limit=80)}


def _comments_for_fact_batch(facts: list[dict[str, Any]], comments_by_id: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    comments: list[dict[str, Any]] = []
    for fact in facts:
        comment = comments_by_id.get(_clean_text(fact.get("comment_id"), limit=80))
        if comment:
            comments.append(comment)
    return comments


def _analyze_fact_batches(
    *,
    model_name: str,
    comments: list[dict[str, Any]],
    facts: list[dict[str, Any]],
    active_env: dict[str, str],
    hermes_command: str,
    hermes_debug_dir: Path,
    progress_path: Path,
    progress_message: str,
    metrics: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    batch_model, _aggregate_model = _runtime_models(active_env)
    target_bytes = _int_env(active_env, "HERMES_BATCH_TARGET_BYTES", DEFAULT_BATCH_TARGET_BYTES, minimum=1_000)
    facts_batches = _batch_items_by_prompt_budget(
        facts,
        target_bytes=target_bytes,
        prompt_builder=lambda index, total, batch: _build_fact_batch_prompt(
            model_name=model_name,
            batch_index=index,
            total_batches=total,
            facts=batch,
        ),
    )
    if not facts_batches:
        facts_batches = [[]]
    total_batches = max(len(facts_batches), 1)
    metrics["batch_count"] = total_batches
    metrics["batch_target_bytes"] = target_bytes
    metrics["batch_prompt_max_bytes"] = max(
        len(_build_fact_batch_prompt(model_name=model_name, batch_index=index, total_batches=total_batches, facts=batch).encode("utf-8"))
        for index, batch in enumerate(facts_batches, start=1)
    )
    metrics["batch_prompt_total_bytes"] = sum(
        len(_build_fact_batch_prompt(model_name=model_name, batch_index=index, total_batches=total_batches, facts=batch).encode("utf-8"))
        for index, batch in enumerate(facts_batches, start=1)
    )
    comments_lookup = _comments_by_id(comments)
    concurrency = min(_int_env(active_env, "HERMES_BATCH_CONCURRENCY", DEFAULT_BATCH_CONCURRENCY), total_batches)
    metrics["batch_concurrency"] = concurrency
    batch_payloads: list[dict[str, Any] | None] = [None] * total_batches
    batch_fallbacks: list[dict[str, Any]] = []
    batch_wall_started_at = time.monotonic()

    def run_one(index: int, fact_batch: list[dict[str, Any]]) -> tuple[int, dict[str, Any], dict[str, Any] | None]:
        prompt = _build_fact_batch_prompt(model_name=model_name, batch_index=index, total_batches=total_batches, facts=fact_batch)
        try:
            payload = _call_llm_json(
                prompt,
                hermes_command=hermes_command,
                env=active_env,
                model=batch_model,
                debug_dir=hermes_debug_dir,
                call_label=f"batch_{index:03d}",
                metrics=metrics,
                metric_stage="batch",
            )
            if not isinstance(payload, dict):
                raise ValueError("hermes_invalid_json:batch payload is not object")
            payload.setdefault("batch", f"{index}/{total_batches}")
            return index, payload, None
        except Exception as exc:
            reason = str(exc) or exc.__class__.__name__
            fallback = {"batch": index, "reason": reason}
            comment_batch = _comments_for_fact_batch(fact_batch, comments_lookup)
            payload = _local_batch_payload(batch_index=index, total_batches=total_batches, comments=comment_batch, reason=reason)
            return index, payload, fallback

    completed_count = 0
    if concurrency <= 1:
        for index, fact_batch in enumerate(facts_batches, start=1):
            _write_progress(progress_path, percent=10 + int(index * 50 / total_batches), message=f"{progress_message} {index}/{total_batches}")
            result_index, payload, fallback = run_one(index, fact_batch)
            batch_payloads[result_index - 1] = payload
            if fallback:
                batch_fallbacks.append(fallback)
                _write_progress(progress_path, percent=10 + int(index * 50 / total_batches), message=f"{progress_message} {index}/{total_batches} 失败，使用本地批次兜底")
    else:
        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            futures = {
                executor.submit(run_one, index, fact_batch): index
                for index, fact_batch in enumerate(facts_batches, start=1)
            }
            for future in as_completed(futures):
                result_index, payload, fallback = future.result()
                batch_payloads[result_index - 1] = payload
                completed_count += 1
                if fallback:
                    batch_fallbacks.append(fallback)
                _write_progress(
                    progress_path,
                    percent=10 + int(completed_count * 50 / total_batches),
                    message=f"{progress_message} {completed_count}/{total_batches}",
                )

    ordered_payloads = [payload for payload in batch_payloads if isinstance(payload, dict)]
    batch_fallbacks.sort(key=lambda item: int(item.get("batch") or 0))
    metrics["fallbacks"]["batch"] = batch_fallbacks
    _set_llm_wall_duration(metrics, "batch", batch_wall_started_at)
    return ordered_payloads, batch_fallbacks


def generate_outputs(
    *,
    autohome_input: str | Path,
    dcd_input: str | Path | None,
    postprocess_input: str | Path | None,
    summary_output: str | Path,
    terms_output: str | Path,
    wordcloud_output_dir: str | Path,
    final_report_output: str | Path,
    qa_chunks_output: str | Path,
    model_name: str,
    progress_file: str | Path,
    summary_script: str | Path,
    wordcloud_script: str | Path,
    hermes_command: str = "hermes",
    single_platform: bool = False,
    font_path: str | None = None,
    env: dict[str, str] | None = None,
) -> dict[str, Any]:
    active_env = dict(env or os.environ)
    autohome_path = Path(autohome_input)
    dcd_path = Path(dcd_input) if dcd_input else None
    summary_path = Path(summary_output)
    terms_path = Path(terms_output)
    wordcloud_dir = Path(wordcloud_output_dir)
    final_report_path = Path(final_report_output)
    qa_chunks_path = Path(qa_chunks_output)
    normalized_comments_path = final_report_path.parent / "normalized_comments.jsonl"
    analysis_facts_path = final_report_path.parent / "analysis_facts.jsonl"
    llm_metrics_path = final_report_path.parent / "llm_metrics.json"
    progress_path = Path(progress_file)
    debug_dir_value = (active_env.get("HERMES_DEBUG_DIR") or "").strip()
    hermes_debug_dir = Path(debug_dir_value) if debug_dir_value else progress_path.parent.parent / "logs" / "hermes"
    _reset_hermes_debug_dir(hermes_debug_dir)
    comments = extract_whitelisted_comments(autohome_input=autohome_path, dcd_input=dcd_path, model_name=model_name)
    _write_normalized_comments_jsonl(normalized_comments_path, comments)
    analysis_facts = _build_analysis_facts(comments)
    _write_analysis_facts_jsonl(analysis_facts_path, analysis_facts)
    _write_progress(progress_path, percent=10, message=f"读取脱敏原评论 {len(comments)} 条")
    batch_model, aggregate_model = _runtime_models(active_env)
    llm_metrics = _new_llm_metrics(
        mode=_hermes_llm_mode(active_env),
        batch_model=batch_model,
        aggregate_model=aggregate_model,
        comment_count=len(comments),
        fact_count=len(analysis_facts),
    )
    metrics_started_at = time.monotonic()

    fallback_reason = ""
    api_key = (active_env.get("LLM_API_KEY") or active_env.get("DEEPSEEK_API_KEY") or "").strip()
    if not api_key:
        fallback_reason = "hermes_disabled:missing_llm_api_key"

    if not fallback_reason:
        try:
            batch_payloads, batch_fallbacks = _analyze_fact_batches(
                model_name=model_name,
                comments=comments,
                facts=analysis_facts,
                active_env=active_env,
                hermes_command=hermes_command,
                hermes_debug_dir=hermes_debug_dir,
                progress_path=progress_path,
                progress_message="Hermes 分析批次",
                metrics=llm_metrics,
            )

            _write_progress(progress_path, percent=70, message="Hermes 汇总批次结果")
            aggregate_source = _source_name(active_env, batch_fallbacks=batch_fallbacks)
            aggregate_fallback_reason = ""
            aggregate_env = dict(active_env)
            aggregate_timeout = (active_env.get("HERMES_AGGREGATE_TIMEOUT_SECONDS") or "").strip()
            if aggregate_timeout:
                aggregate_env["HERMES_TIMEOUT_SECONDS"] = aggregate_timeout
            aggregate_prompt = _build_aggregate_prompt(model_name=model_name, comments=comments, batch_payloads=batch_payloads)
            try:
                aggregate_payload = _call_aggregate_llm_json(
                    aggregate_prompt,
                    hermes_command=hermes_command,
                    env=aggregate_env,
                    model=aggregate_model,
                    debug_dir=hermes_debug_dir,
                    call_label="aggregate",
                    metrics=llm_metrics,
                    metric_stage="aggregate",
                )
                if not isinstance(aggregate_payload, dict):
                    raise ValueError("hermes_invalid_json:aggregate payload is not object")
            except Exception as exc:
                aggregate_fallback_reason = str(exc)
                aggregate_source = _source_name(active_env, batch_fallbacks=batch_fallbacks, aggregate_local=True)
                llm_metrics["fallbacks"]["aggregate"] = aggregate_fallback_reason
                _write_progress(progress_path, percent=82, message="Hermes 汇总超时，使用批次结果本地归并")
                aggregate_payload = _local_aggregate_payload(model_name=model_name, comments=comments, batch_payloads=batch_payloads)

            normalized = _normalize_hermes_payload(aggregate_payload, model_name=model_name, comments=comments, batch_payloads=batch_payloads)
            _write_summary_workbook(summary_path, model_name=model_name, comments=comments, result=normalized)
            _write_terms_workbook(terms_path, normalized.keyword_rankings)
            image_paths = _write_wordclouds(wordcloud_dir, model_name=model_name, rankings=normalized.keyword_rankings, font_path=font_path)
            _write_report_json(final_report_path, normalized.report)
            _write_qa_chunks(qa_chunks_path, normalized.qa_chunks)
            _write_validation(summary_path.with_suffix(".validation.json"), source=aggregate_source, degraded=False)
            llm_metrics["source"] = aggregate_source
            llm_metrics["durations_ms"]["total"] = int((time.monotonic() - metrics_started_at) * 1000)
            _set_llm_wall_duration(llm_metrics, "total", metrics_started_at)
            _write_llm_metrics(llm_metrics_path, llm_metrics)
            _write_progress(progress_path, percent=100, message="Hermes 输出已生成")
            result = {
                "status": "success",
                "degraded": False,
                "source": aggregate_source,
                "summary_path": str(summary_path),
                "terms_path": str(terms_path),
                "final_report_path": str(final_report_path),
                "qa_chunks_path": str(qa_chunks_path),
                "normalized_comments_path": str(normalized_comments_path),
                "analysis_facts_path": str(analysis_facts_path),
                "llm_metrics_path": str(llm_metrics_path),
                "image_paths": image_paths,
                "postprocess_path": str(postprocess_input or ""),
            }
            if aggregate_fallback_reason:
                result["aggregate_fallback_reason"] = aggregate_fallback_reason
            if batch_fallbacks:
                result["batch_fallbacks"] = batch_fallbacks
            return result
        except Exception as exc:
            fallback_reason = str(exc)

    _write_progress(progress_path, percent=75, message="Hermes 不可用，切换到规则兜底", degraded=True)
    llm_metrics["source"] = "rule-fallback"
    llm_metrics["fallback_reason"] = fallback_reason
    llm_metrics["durations_ms"]["total"] = int((time.monotonic() - metrics_started_at) * 1000)
    _set_llm_wall_duration(llm_metrics, "total", metrics_started_at)
    _write_llm_metrics(llm_metrics_path, llm_metrics)
    fallback = _run_rule_fallback(
        autohome_input=autohome_path,
        dcd_input=dcd_path,
        summary_output=summary_path,
        terms_output=terms_path,
        wordcloud_output_dir=wordcloud_dir,
        final_report_output=final_report_path,
        qa_chunks_output=qa_chunks_path,
        model_name=model_name,
        progress_file=progress_path,
        summary_script=Path(summary_script),
        wordcloud_script=Path(wordcloud_script),
        single_platform=single_platform,
        font_path=font_path,
    )
    _write_progress(progress_path, percent=100, message="规则兜底输出已生成", degraded=True)
    return {
        "status": "degraded",
        "degraded": True,
        "source": "rule-fallback",
        "fallback_reason": fallback_reason,
        "normalized_comments_path": str(normalized_comments_path),
        "analysis_facts_path": str(analysis_facts_path),
        "llm_metrics_path": str(llm_metrics_path),
        **fallback,
    }


def generate_time_report_outputs(
    *,
    autohome_input: str | Path,
    dcd_input: str | Path | None,
    output_dir: str | Path,
    model_name: str,
    start_date: str,
    end_date: str,
    hermes_command: str = "hermes",
    font_path: str | None = None,
    env: dict[str, str] | None = None,
    progress_file: str | Path | None = None,
) -> dict[str, Any]:
    active_env = dict(env or os.environ)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    progress_path = Path(progress_file) if progress_file else output_path / "time_report.progress.json"
    debug_dir_value = (active_env.get("HERMES_DEBUG_DIR") or "").strip()
    hermes_debug_dir = Path(debug_dir_value) if debug_dir_value else output_path / "logs" / "hermes"
    _reset_hermes_debug_dir(hermes_debug_dir)

    comments = extract_whitelisted_comments(autohome_input=autohome_input, dcd_input=dcd_input, model_name=model_name)
    selected_comments = _filter_comments_for_date_range(comments, start_date=start_date, end_date=end_date)
    if not selected_comments:
        raise ValueError("no_comments_in_date_range")

    platform_counts = dict(Counter(comment["platform"] for comment in selected_comments))
    _write_progress(progress_path, percent=10, message=f"读取时间范围内脱敏原评论 {len(selected_comments)} 条")

    date_label = f"{start_date}_{end_date}"
    summary_path = output_path / f"{model_name}_{date_label}_时间范围口碑摘要.xlsx"
    terms_path = output_path / f"{model_name}_{date_label}_词云词项清单.xlsx"
    final_report_path = output_path / "final_report.json"
    qa_chunks_path = output_path / "qa_chunks.json"
    normalized_comments_path = output_path / "normalized_comments.jsonl"
    analysis_facts_path = output_path / "analysis_facts.jsonl"
    llm_metrics_path = output_path / "llm_metrics.json"
    validation_path = summary_path.with_suffix(".validation.json")
    _write_normalized_comments_jsonl(normalized_comments_path, selected_comments)
    analysis_facts = _build_analysis_facts(selected_comments)
    _write_analysis_facts_jsonl(analysis_facts_path, analysis_facts)
    batch_model, aggregate_model = _runtime_models(active_env)
    llm_metrics = _new_llm_metrics(
        mode=_hermes_llm_mode(active_env),
        batch_model=batch_model,
        aggregate_model=aggregate_model,
        comment_count=len(selected_comments),
        fact_count=len(analysis_facts),
    )
    metrics_started_at = time.monotonic()

    api_key = (active_env.get("LLM_API_KEY") or active_env.get("DEEPSEEK_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError("hermes_disabled:missing_llm_api_key")

    batch_payloads, batch_fallbacks = _analyze_fact_batches(
        model_name=model_name,
        comments=selected_comments,
        facts=analysis_facts,
        active_env=active_env,
        hermes_command=hermes_command,
        hermes_debug_dir=hermes_debug_dir,
        progress_path=progress_path,
        progress_message="Hermes 分析时间范围批次",
        metrics=llm_metrics,
    )

    _write_progress(progress_path, percent=70, message="Hermes 汇总时间范围批次结果")
    aggregate_source = _source_name(active_env, batch_fallbacks=batch_fallbacks)
    aggregate_fallback_reason = ""
    aggregate_env = dict(active_env)
    aggregate_timeout = (active_env.get("HERMES_AGGREGATE_TIMEOUT_SECONDS") or "").strip()
    if aggregate_timeout:
        aggregate_env["HERMES_TIMEOUT_SECONDS"] = aggregate_timeout
    aggregate_prompt = _build_aggregate_prompt(model_name=model_name, comments=selected_comments, batch_payloads=batch_payloads)
    try:
        aggregate_payload = _call_aggregate_llm_json(
            aggregate_prompt,
            hermes_command=hermes_command,
            env=aggregate_env,
            model=aggregate_model,
            debug_dir=hermes_debug_dir,
            call_label="aggregate",
            metrics=llm_metrics,
            metric_stage="aggregate",
        )
        if not isinstance(aggregate_payload, dict):
            raise ValueError("hermes_invalid_json:aggregate payload is not object")
    except Exception as exc:
        aggregate_fallback_reason = str(exc)
        aggregate_source = _source_name(active_env, batch_fallbacks=batch_fallbacks, aggregate_local=True)
        llm_metrics["fallbacks"]["aggregate"] = aggregate_fallback_reason
        _write_progress(progress_path, percent=82, message="Hermes 汇总超时，使用时间范围批次结果本地归并")
        aggregate_payload = _local_aggregate_payload(model_name=model_name, comments=selected_comments, batch_payloads=batch_payloads)

    normalized = _normalize_hermes_payload(aggregate_payload, model_name=model_name, comments=selected_comments, batch_payloads=batch_payloads)
    normalized.report["report_type"] = "time_range"
    normalized.report["time_range"] = {"start_date": start_date, "end_date": end_date}
    normalized.report["sample_count"] = len(selected_comments)
    normalized.report["platform_counts"] = platform_counts

    _write_summary_workbook(summary_path, model_name=model_name, comments=selected_comments, result=normalized)
    _write_terms_workbook(terms_path, normalized.keyword_rankings)
    image_paths = _write_wordclouds(output_path, model_name=model_name, rankings=normalized.keyword_rankings, font_path=font_path)
    _write_report_json(final_report_path, normalized.report)
    _write_qa_chunks(qa_chunks_path, normalized.qa_chunks)
    _write_validation(validation_path, source=aggregate_source, degraded=False)
    llm_metrics["source"] = aggregate_source
    llm_metrics["durations_ms"]["total"] = int((time.monotonic() - metrics_started_at) * 1000)
    _set_llm_wall_duration(llm_metrics, "total", metrics_started_at)
    _write_llm_metrics(llm_metrics_path, llm_metrics)
    _write_progress(progress_path, percent=100, message="时间范围 Hermes 一页纸已生成")

    artifact_paths = [
        str(final_report_path),
        str(normalized_comments_path),
        str(analysis_facts_path),
        str(llm_metrics_path),
        str(summary_path),
        str(validation_path),
        str(terms_path),
        str(qa_chunks_path),
        *image_paths,
    ]
    result = {
        "status": "completed",
        "degraded": False,
        "source": aggregate_source,
        "sample_count": len(selected_comments),
        "platform_counts": platform_counts,
        "summary_path": str(summary_path),
        "terms_path": str(terms_path),
        "final_report_path": str(final_report_path),
        "qa_chunks_path": str(qa_chunks_path),
        "normalized_comments_path": str(normalized_comments_path),
        "analysis_facts_path": str(analysis_facts_path),
        "llm_metrics_path": str(llm_metrics_path),
        "validation_path": str(validation_path),
        "image_paths": image_paths,
        "artifact_paths": artifact_paths,
        "report_json": normalized.report,
    }
    if aggregate_fallback_reason:
        result["aggregate_fallback_reason"] = aggregate_fallback_reason
    if batch_fallbacks:
        result["batch_fallbacks"] = batch_fallbacks
    return result


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate final koubei outputs through Hermes Agent with rule fallback.")
    parser.add_argument("--autohome-input", required=True)
    parser.add_argument("--dcd-input")
    parser.add_argument("--postprocess-input")
    parser.add_argument("--summary-output", required=True)
    parser.add_argument("--terms-output", required=True)
    parser.add_argument("--wordcloud-output-dir", required=True)
    parser.add_argument("--final-report-output", required=True)
    parser.add_argument("--qa-chunks-output", required=True)
    parser.add_argument("--model-name", required=True)
    parser.add_argument("--progress-file", required=True)
    parser.add_argument("--summary-script", required=True)
    parser.add_argument("--wordcloud-script", required=True)
    parser.add_argument("--hermes-command", default=os.getenv("HERMES_COMMAND", "hermes"))
    parser.add_argument("--font-path")
    parser.add_argument("--single-platform", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv or sys.argv[1:])
    result = generate_outputs(
        autohome_input=args.autohome_input,
        dcd_input=args.dcd_input,
        postprocess_input=args.postprocess_input,
        summary_output=args.summary_output,
        terms_output=args.terms_output,
        wordcloud_output_dir=args.wordcloud_output_dir,
        final_report_output=args.final_report_output,
        qa_chunks_output=args.qa_chunks_output,
        model_name=args.model_name,
        progress_file=args.progress_file,
        summary_script=args.summary_script,
        wordcloud_script=args.wordcloud_script,
        hermes_command=args.hermes_command,
        single_platform=args.single_platform,
        font_path=args.font_path,
    )
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
