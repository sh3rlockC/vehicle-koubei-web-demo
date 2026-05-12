from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import Settings
from app.models import Job, JobArtifact, JobTimeReport


PLATFORM_AUTOHOME = "汽车之家"
PLATFORM_DCD = "懂车帝"


@dataclass(frozen=True)
class CommentRecord:
    comment_id: str
    platform: str
    date: str
    date_value: date | None
    model_name: str
    positive_text: str
    negative_text: str
    full_text: str

    def public_dict(self) -> dict[str, str]:
        return {
            "comment_id": self.comment_id,
            "platform": self.platform,
            "date": self.date,
            "model_name": self.model_name,
            "positive_text": self.positive_text,
            "negative_text": self.negative_text,
            "full_text": self.full_text,
        }


def clean_text(value: object, *, limit: int = 900) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def parse_comment_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value, limit=80)
    if not text:
        return None
    match = re.search(r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def parse_date_param(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def first_value(row: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = clean_text(row.get(key))
        if value:
            return value
    return ""


def iter_sheet_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            raw_rows = list(worksheet.iter_rows(values_only=True))
            if not raw_rows:
                continue
            header = [clean_text(cell) for cell in raw_rows[0]]
            if not any(header):
                continue
            for raw_row in raw_rows[1:]:
                row = {header[index]: raw_row[index] for index in range(min(len(header), len(raw_row))) if header[index]}
                if any(clean_text(value) for value in row.values()):
                    rows.append(row)
    finally:
        workbook.close()
    return rows


def comment_from_row(row: dict[str, Any], *, platform: str, model_name: str, comment_id: str) -> CommentRecord | None:
    positive_text = first_value(row, ("最满意", "满意", "优点", "优势", "正向反馈"))
    negative_text = first_value(row, ("最不满意", "不满意", "缺点", "槽点", "负向反馈"))
    full_text = first_value(row, ("评价详情", "评价全文", "口碑内容", "内容", "正文", "评论", "原文"))
    raw_date = first_value(row, ("发表日期", "发布时间", "日期", "时间"))
    parsed_date = parse_comment_date(raw_date)
    row_model = first_value(row, ("车型", "评价车型", "车款", "车系", "车型名称")) or model_name
    if not any([positive_text, negative_text, full_text]):
        return None
    return CommentRecord(
        comment_id=comment_id,
        platform=platform,
        date=parsed_date.isoformat() if parsed_date else "",
        date_value=parsed_date,
        model_name=row_model,
        positive_text=positive_text,
        negative_text=negative_text,
        full_text=full_text,
    )


def _artifact_path_by_stage(artifacts: list[JobArtifact], stage: str) -> Path | None:
    for artifact in artifacts:
        if artifact.source_stage == stage and artifact.artifact_path.lower().endswith(".xlsx"):
            path = Path(artifact.artifact_path)
            if path.exists():
                return path
    return None


def raw_comment_paths(db: Session, settings: Settings, job: Job) -> tuple[Path | None, Path | None]:
    artifacts = (
        db.query(JobArtifact)
        .filter(JobArtifact.job_id == job.job_id)
        .order_by(JobArtifact.id.asc())
        .all()
    )
    autohome = _artifact_path_by_stage(artifacts, "collecting_autohome")
    dcd = _artifact_path_by_stage(artifacts, "collecting_dcd")
    raw_dir = settings.artifact_root_path / job.job_id / "outputs" / "raw"
    return (
        autohome or raw_dir / f"ZJ{job.model_name}原始口碑.xlsx",
        dcd or raw_dir / f"DCD口碑_{job.model_name}.xlsx",
    )


def extract_job_comments(db: Session, settings: Settings, job: Job) -> list[CommentRecord]:
    autohome_path, dcd_path = raw_comment_paths(db, settings, job)
    comments: list[CommentRecord] = []
    if autohome_path and autohome_path.exists():
        index = 1
        for row in iter_sheet_rows(autohome_path):
            comment = comment_from_row(row, platform=PLATFORM_AUTOHOME, model_name=job.model_name, comment_id=f"autohome_{index:04d}")
            if comment:
                comments.append(comment)
                index += 1
    if dcd_path and dcd_path.exists():
        index = 1
        for row in iter_sheet_rows(dcd_path):
            comment = comment_from_row(row, platform=PLATFORM_DCD, model_name=job.model_name, comment_id=f"dcd_{index:04d}")
            if comment:
                comments.append(comment)
                index += 1
    return comments


def filter_comments_by_date(comments: list[CommentRecord], *, start_date: str, end_date: str) -> list[CommentRecord]:
    start = parse_date_param(start_date)
    end = parse_date_param(end_date)
    if start > end:
        return []
    filtered = [comment for comment in comments if comment.date_value is not None and start <= comment.date_value <= end]
    return sorted(filtered, key=lambda comment: (comment.date_value or date.min, comment.platform, comment.comment_id))


def comment_summary(job_id: str, comments: list[CommentRecord]) -> dict[str, Any]:
    dated = [comment for comment in comments if comment.date_value is not None]
    daily_counts = Counter(comment.date for comment in dated)
    platform_counts = Counter(comment.platform for comment in comments)
    dates = sorted(daily_counts)
    return {
        "job_id": job_id,
        "total_count": len(comments),
        "dated_count": len(dated),
        "undated_count": len(comments) - len(dated),
        "date_min": dates[0] if dates else None,
        "date_max": dates[-1] if dates else None,
        "daily_counts": [{"date": value, "count": daily_counts[value]} for value in dates],
        "platform_counts": dict(platform_counts),
    }


def platform_counts(comments: list[CommentRecord]) -> dict[str, int]:
    return dict(Counter(comment.platform for comment in comments))


def time_report_artifacts(report: JobTimeReport) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []
    for raw_path in report.artifact_paths or []:
        path = Path(str(raw_path))
        lower = path.name.lower()
        if lower.endswith("final_report.json"):
            artifact_type = "final_report_json"
        elif lower.endswith("_词云词项清单.xlsx"):
            artifact_type = "wordcloud_terms_excel"
        elif lower.endswith(".xlsx"):
            artifact_type = "summary_excel"
        elif lower.endswith(".png"):
            artifact_type = "wordcloud_png"
        elif lower.endswith(".json"):
            artifact_type = "json"
        else:
            artifact_type = "file"
        items.append({"name": path.name, "path": str(path), "type": artifact_type})
    return items


def time_report_payload(report: JobTimeReport) -> dict[str, Any]:
    return {
        "report_id": report.report_id,
        "job_id": report.job_id,
        "model_name": report.model_name,
        "date_range": {"start_date": report.start_date, "end_date": report.end_date},
        "status": report.status,
        "sample_count": report.sample_count,
        "platform_counts": report.platform_counts or {},
        "source": report.source,
        "report_json": report.report_json or None,
        "artifacts": time_report_artifacts(report),
        "zip_url": f"/api/jobs/{report.job_id}/time-reports/{report.report_id}/artifacts.zip",
        "error_code": report.error_code,
        "error_message": report.error_message,
        "queue_job_id": report.queue_job_id,
        "created_at": report.created_at,
        "updated_at": report.updated_at,
        "completed_at": report.completed_at,
    }
