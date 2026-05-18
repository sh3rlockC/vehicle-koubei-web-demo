from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

KeywordDirection = Literal["positive", "negative"]
logger = logging.getLogger(__name__)

OUTER_KEYWORD_WRAPPERS = (
    ("「", "」"),
    ("『", "』"),
    ("“", "”"),
    ("‘", "’"),
    ("《", "》"),
    ("〈", "〉"),
    ("【", "】"),
    ("（", "）"),
    ("(", ")"),
    ("[", "]"),
    ('"', '"'),
    ("'", "'"),
)


def _empty_keyword_rankings() -> dict[str, list[dict[str, int | str]]]:
    return {"positive": [], "negative": [], "combined": []}


def _normalize_keyword_term(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"[\u200b-\u200f\ufeff]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    while len(text) >= 2:
        for left, right in OUTER_KEYWORD_WRAPPERS:
            if text.startswith(left) and text.endswith(right):
                inner = text[len(left) : len(text) - len(right)].strip()
                if inner:
                    text = inner
                    break
        else:
            break
    return text


def _iter_table_rows(worksheet, *, limit: int | None = None) -> list[dict[str, str]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        values = {header[index]: row[index] for index in range(min(len(header), len(row))) if header[index]}
        if not any(value not in (None, "") for value in values.values()):
            continue
        normalized = {key: "" if value is None else str(value) for key, value in values.items()}
        data_rows.append(normalized)
        if limit is not None and len(data_rows) >= limit:
            break
    return data_rows


def _parse_sample_counts(overview_rows: list[dict[str, str]]) -> dict[str, int]:
    for row in overview_rows:
        if row.get("模块") != "平台样本":
            continue
        content = row.get("内容", "")
        match = re.search(r"汽车之家\s+(\d+)\s+条.*?懂车帝\s+(\d+)\s+条", content)
        if match:
            return {
                "autohome_count": int(match.group(1)),
                "dcd_count": int(match.group(2)),
            }
    return {"autohome_count": 0, "dcd_count": 0}


def _read_one_pager_lines(worksheet, *, limit: int = 12) -> list[str]:
    lines: list[str] = []
    for row in worksheet.iter_rows(values_only=True):
        values = [str(value).strip() for value in row if value not in (None, "")]
        if not values:
            continue
        lines.append(" ".join(values))
        if len(lines) >= limit:
            break
    return lines


def _numeric_weight(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _count_from_weight(value: float) -> int:
    return int(round(value))


def _sort_keyword_rows(rows: list[tuple[str, float, int]], *, limit: int) -> list[dict[str, int | str]]:
    sorted_rows = sorted(rows, key=lambda item: (-item[1], item[2], item[0]))
    return [{"term": term, "count": _count_from_weight(weight)} for term, weight, _index in sorted_rows[:limit]]


def _rankings_from_breakdown(worksheet, *, limit: int) -> dict[str, list[dict[str, int | str]]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return _empty_keyword_rankings()

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    direction_weights: dict[KeywordDirection, dict[str, tuple[float, int]]] = {"positive": {}, "negative": {}}
    combined_weights: dict[str, tuple[float, int]] = {}

    for index, row in enumerate(rows[1:]):
        values = {header[cell_index]: row[cell_index] for cell_index in range(min(len(header), len(row))) if header[cell_index]}
        direction = values.get("direction")
        term = _normalize_keyword_term(values.get("term"))
        weight = _numeric_weight(values.get("weight"))
        if direction not in {"positive", "negative"} or not term or weight <= 0:
            continue

        direction_key: KeywordDirection = "positive" if direction == "positive" else "negative"
        previous_direction_weight, first_direction_index = direction_weights[direction_key].get(term, (0.0, index))
        direction_weights[direction_key][term] = (
            previous_direction_weight + weight,
            min(first_direction_index, index),
        )
        previous_weight, first_index = combined_weights.get(term, (0.0, index))
        combined_weights[term] = (previous_weight + weight, min(first_index, index))

    positive_rows = [(term, weight, index) for term, (weight, index) in direction_weights["positive"].items()]
    negative_rows = [(term, weight, index) for term, (weight, index) in direction_weights["negative"].items()]
    combined_rows = [(term, weight, index) for term, (weight, index) in combined_weights.items()]
    return {
        "positive": _sort_keyword_rows(positive_rows, limit=limit),
        "negative": _sort_keyword_rows(negative_rows, limit=limit),
        "combined": _sort_keyword_rows(combined_rows, limit=limit),
    }


def _rankings_from_terms_sheets(workbook, *, limit: int) -> dict[str, list[dict[str, int | str]]]:
    direction_to_sheet: dict[KeywordDirection, str] = {
        "positive": "positive_terms",
        "negative": "negative_terms",
    }
    grouped: dict[KeywordDirection, dict[str, tuple[float, int]]] = {"positive": {}, "negative": {}}
    combined_weights: dict[str, tuple[float, int]] = {}
    global_index = 0

    for direction, sheet_name in direction_to_sheet.items():
        if sheet_name not in workbook.sheetnames:
            continue
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        for row in rows[1:]:
            global_index += 1
            values = {header[cell_index]: row[cell_index] for cell_index in range(min(len(header), len(row))) if header[cell_index]}
            term = _normalize_keyword_term(values.get("term"))
            weight = _numeric_weight(values.get("weight"))
            if not term or weight <= 0:
                continue

            previous_weight, first_index = grouped[direction].get(term, (0.0, global_index))
            grouped[direction][term] = (previous_weight + weight, min(first_index, global_index))
            combined_weight, combined_index = combined_weights.get(term, (0.0, global_index))
            combined_weights[term] = (combined_weight + weight, min(combined_index, global_index))

    positive_rows = [(term, weight, index) for term, (weight, index) in grouped["positive"].items()]
    negative_rows = [(term, weight, index) for term, (weight, index) in grouped["negative"].items()]
    combined_rows = [(term, weight, index) for term, (weight, index) in combined_weights.items()]
    return {
        "positive": _sort_keyword_rows(positive_rows, limit=limit),
        "negative": _sort_keyword_rows(negative_rows, limit=limit),
        "combined": _sort_keyword_rows(combined_rows, limit=limit),
    }


def read_wordcloud_terms_workbook(path: str | Path, *, limit: int = 10) -> dict[str, list[dict[str, int | str]]]:
    workbook = load_workbook(Path(path), data_only=True)
    if "platform_breakdown" in workbook.sheetnames:
        rankings = _rankings_from_breakdown(workbook["platform_breakdown"], limit=limit)
        if rankings["positive"] or rankings["negative"] or rankings["combined"]:
            return rankings
    return _rankings_from_terms_sheets(workbook, limit=limit)


def read_wordcloud_terms_workbook_or_empty(path: str | Path, *, limit: int = 10) -> dict[str, list[dict[str, int | str]]]:
    try:
        return read_wordcloud_terms_workbook(path, limit=limit)
    except Exception as exc:
        logger.warning("failed to read wordcloud terms workbook %s: %s", path, exc)
        return _empty_keyword_rankings()


def read_summary_workbook(path: str | Path) -> dict:
    workbook = load_workbook(Path(path), data_only=True)

    overview_rows = _iter_table_rows(workbook["总览摘要"])
    compare_rows = _iter_table_rows(workbook["跨平台对比"], limit=8)
    business_rows = _iter_table_rows(workbook["综合业务摘要"])
    opportunity_rows = _iter_table_rows(workbook["产品机会点"], limit=8)
    one_pager_lines = _read_one_pager_lines(workbook["一页纸总结"])

    return {
        "sample_counts": _parse_sample_counts(overview_rows),
        "overview_rows": overview_rows,
        "compare_rows": compare_rows,
        "business_rows": business_rows,
        "opportunity_rows": opportunity_rows,
        "one_pager_lines": one_pager_lines,
    }
